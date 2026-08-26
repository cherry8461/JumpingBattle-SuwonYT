import os
import re
import time
import sys
import json
import sqlite3
import threading
import gspread
import imaplib
import email
import base64
import logging
import hmac
from functools import wraps
from io import BytesIO
from dotenv import load_dotenv
from email.header import decode_header
from datetime import datetime, timedelta

# Flask 및 SocketIO (가장 표준적인 형태)
from flask import Flask, render_template, render_template_string, request, redirect, url_for, jsonify, send_file, session
from flask_socketio import SocketIO, emit

# 엑셀 및 기타
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from oauth2client.service_account import ServiceAccountCredentials
from gspread_formatting import format_cell_ranges, Color, set_column_widths
from monitor_core.database import get_db_connection
from monitor_core.settings import (
    DATA_DIR,
    DB_FILE,
    LOG_DIR,
    PROJECT_ROOT,
    SERVER_HOST,
    SERVER_LOG_DIR,
    SERVER_PORT,
)
from monitor_modules.naver_reservations import create_naver_reservations_blueprint

# ========================================================
# Flask
# ========================================================
web = Flask(__name__)
web.config["SECRET_KEY"] = os.getenv("FLASK_SECRET_KEY", "development-only-change-me")
web.config["TEMPLATES_AUTO_RELOAD"] = True
web.jinja_env.auto_reload = True

from werkzeug.middleware.proxy_fix import ProxyFix
web.wsgi_app = ProxyFix(web.wsgi_app, x_proto=1, x_host=1)

socketio = SocketIO(web, cors_allowed_origins="*", ping_timeout=60, ping_interval=15, async_mode='threading')
web.register_blueprint(create_naver_reservations_blueprint(socketio))

OFFSET_FILE = "log_offset.dat"
PAD_COUNT = 4

base_path = str(PROJECT_ROOT)
key_path = os.path.join(base_path, "config", "google_key.json")
gc = None
sh = None 

# 인증 진행
def init_google_sheets():
    global gc, sh
    try:
        base_path = os.path.dirname(os.path.abspath(__file__))
        key_path = os.path.join(base_path, "config", "google_key.json")
        gc = gspread.service_account(filename=key_path)
        sh = gc.open("2026년 월 정산표")
        print("구글 시트 연결 성공!")
    except Exception as e:
        print(f"연결 실패: {e}")

def save_to_google_sheet(data):
    current_month_name = f"{int(datetime.now().strftime('%m'))}월"
    
    try:
        return sh.worksheet(current_month_name)
    except gspread.exceptions.WorksheetNotFound:
        # 1. 시트 생성 (생성 직후에는 요청 간격을 아주 약간 둠)
        worksheet = sh.add_worksheet(title=current_month_name, rows=50, cols=25, index=0)
        time.sleep(0.5)
        sheet_id = worksheet.id
        
        now = datetime.now()
        year_short = now.strftime('%y')
        sheet_id = worksheet.id

        color_yellow = {"red": 1.0, "green": 1.0, "blue": 0.0}
        color_soft_blue = {"red": 0.706, "green": 0.776, "blue": 0.906}
        color_title_bg = {"red": 0.9, "green": 0.92, "blue": 0.95}
        color_light_yellow = {"red": 1.0, "green": 0.898, "blue": 0.6}
        color_pink = {"red": 0.95, "green": 0.8, "blue": 0.8}
        color_mint = {"red": 0.6, "green": 0.8, "blue": 0.8}
        color_orange = {"red": 1.0, "green": 0.8, "blue": 0.6}
        color_sky = {"red": 0.6, "green": 0.8, "blue": 0.9}
        color_light_green = {"red": 0.714, "green": 0.843, "blue": 0.659}
        color_lavender = {"red": 0.706, "green": 0.655, "blue": 0.839}
        color_header_bg = {"red": 0.95, "green": 0.95, "blue": 0.95}

        def get_range(range_str):
            parts = re.split(r'[:!]', range_str)
            start_cell = parts[0]
            end_cell = parts[1] if len(parts) > 1 else start_cell
            
            def cell_to_idx(cell):
                col_part = re.findall(r'[A-Z]+', cell)[0]
                # 열 이름(A, B, AA 등)을 인덱스 숫자로 변환
                col = 0
                for char in col_part:
                    col = col * 26 + (ord(char) - ord('A') + 1)
                row = int(re.findall(r'\d+', cell)[0]) - 1
                return row, col - 1

            s_row, s_col = cell_to_idx(start_cell)
            e_row, e_col = cell_to_idx(end_cell)
            return {
                "sheetId": sheet_id,
                "startRowIndex": s_row, "endRowIndex": e_row + 1,
                "startColumnIndex": s_col, "endColumnIndex": e_col + 1
            }

        # --- 2. 데이터 통합 입력 (배치 업데이트) ---
        # 개별 update_acell 대신 범위를 지정해서 한꺼번에 밀어넣습니다.
        input_data = [
            {"range": "B2", "values": [[f"{year_short}년 {current_month_name} "]]},
            {"range": "D2", "values": [["월매출"]]},
            {"range": "E2", "values": [["=SUM(P5:P50)"]]}, # 넉넉히 50행까지
            {"range": "G2", "values": [["카드"]]},
            {"range": "H2", "values": [["=SUM(L5:L50)"]]},
            {"range": "I2", "values": [["현금"]]},
            {"range": "J2", "values": [["=SUM(M5:M50)"]]},
            {"range": "L2", "values": [["계좌"]]},
            {"range": "M2", "values": [["=SUM(N5:N50)"]]},
            {"range": "O2", "values": [["파티룸"]]},
            {"range": "P2", "values": [["=SUM(U5:U50)"]]},
            {"range": "R2", "values": [["다회권"]]},
            {"range": "T2", "values": [["=SUM(V5:V50)"]]},
            {"range": "D4", "values": [["게임결제", "", "", ""]]}, # 병합용 데이터 채우기
            {"range": "H4", "values": [["기타판매", "", ""]]},
            {"range": "L4", "values": [["전체매출", "", "", "", ""]]},
            {"range": "R4", "values": [["방문통계", ""]]},
            {"range": "U4:V4", "values": [["파티룸", "다회권"]]},
            # 4행 헤더 한꺼번에 입력
            {"range": "B5:V5", "values": [[
                "일자", "요일", "예약금", "카드", "현금", "계좌", 
                "카드", "현금", "계좌", "", "예약금+카드", "현금", "계좌", 
                "사용다회/쿠폰", "총매출", "", "팀", "인원수", "", "파티룸", "다회권"
            ]]}
        ]
        worksheet.batch_update(input_data, value_input_option='USER_ENTERED')

        # --- 3. 모든 서식 및 병합 통합 요청 (sh.batch_update) ---
        requests = [
            # A. 셀 병합 통합 요청
            {"mergeCells": {"range": {"sheetId": sheet_id, "startRowIndex": 1, "endRowIndex": 2, "startColumnIndex": 1, "endColumnIndex": 3}, "mergeType": "MERGE_ALL"}}, # B2:C2
            {"mergeCells": {"range": {"sheetId": sheet_id, "startRowIndex": 1, "endRowIndex": 2, "startColumnIndex": 4, "endColumnIndex": 6}, "mergeType": "MERGE_ALL"}}, # E2:F2
            {"mergeCells": {"range": {"sheetId": sheet_id, "startRowIndex": 1, "endRowIndex": 2, "startColumnIndex": 15, "endColumnIndex": 17}, "mergeType": "MERGE_ALL"}}, # P2:Q2
            {"mergeCells": {"range": {"sheetId": sheet_id, "startRowIndex": 1, "endRowIndex": 2, "startColumnIndex": 17, "endColumnIndex": 19}, "mergeType": "MERGE_ALL"}}, # R2:S2
            {"mergeCells": {"range": {"sheetId": sheet_id, "startRowIndex": 1, "endRowIndex": 2, "startColumnIndex": 19, "endColumnIndex": 21}, "mergeType": "MERGE_ALL"}}, # T2:U2
            {"mergeCells": {"range": {"sheetId": sheet_id, "startRowIndex": 3, "endRowIndex": 4, "startColumnIndex": 3, "endColumnIndex": 7}, "mergeType": "MERGE_ALL"}}, # D3:G3
            {"mergeCells": {"range": {"sheetId": sheet_id, "startRowIndex": 3, "endRowIndex": 4, "startColumnIndex": 7, "endColumnIndex": 10}, "mergeType": "MERGE_ALL"}}, # H3:J3
            {"mergeCells": {"range": {"sheetId": sheet_id, "startRowIndex": 3, "endRowIndex": 4, "startColumnIndex": 11, "endColumnIndex": 16}, "mergeType": "MERGE_ALL"}}, # L3:P3
            {"mergeCells": {"range": {"sheetId": sheet_id, "startRowIndex": 3, "endRowIndex": 4, "startColumnIndex": 17, "endColumnIndex": 19}, "mergeType": "MERGE_ALL"}}, # R3:S3

            # B. 열 너비 통합 설정 (한 번에!)
            {"updateDimensionProperties": {"range": {"sheetId": sheet_id, "dimension": "COLUMNS", "startIndex": 0, "endIndex": 1}, "properties": {"pixelSize": 15}, "fields": "pixelSize"}}, # A
            {"updateDimensionProperties": {"range": {"sheetId": sheet_id, "dimension": "COLUMNS", "startIndex": 1, "endIndex": 3}, "properties": {"pixelSize": 30}, "fields": "pixelSize"}}, # B,C
            {"updateDimensionProperties": {"range": {"sheetId": sheet_id, "dimension": "COLUMNS", "startIndex": 3, "endIndex": 10}, "properties": {"pixelSize": 110}, "fields": "pixelSize"}}, # D,J
            {"updateDimensionProperties": {"range": {"sheetId": sheet_id, "dimension": "COLUMNS", "startIndex": 10, "endIndex": 11}, "properties": {"pixelSize": 15}, "fields": "pixelSize"}}, # K
            {"updateDimensionProperties": {"range": {"sheetId": sheet_id, "dimension": "COLUMNS", "startIndex": 11, "endIndex": 16}, "properties": {"pixelSize": 110}, "fields": "pixelSize"}}, # L,P
            {"updateDimensionProperties": {"range": {"sheetId": sheet_id, "dimension": "COLUMNS", "startIndex": 16, "endIndex": 17}, "properties": {"pixelSize": 15}, "fields": "pixelSize"}}, # Q
            {"updateDimensionProperties": {"range": {"sheetId": sheet_id, "dimension": "COLUMNS", "startIndex": 17, "endIndex": 19}, "properties": {"pixelSize": 50}, "fields": "pixelSize"}}, # R,S
            {"updateDimensionProperties": {"range": {"sheetId": sheet_id, "dimension": "COLUMNS", "startIndex": 19, "endIndex": 20}, "properties": {"pixelSize": 15}, "fields": "pixelSize"}}, # T
        ]
        
        sh.batch_update({"requests": requests})
        time.sleep(0.5)

        fmt_money = {"numberFormat": {"type": "NUMBER", "pattern": "#,##0"}, "horizontalAlignment": "RIGHT"}
        fmt_bold = {"textFormat": {"bold": True}}
        
        design_requests = [
            # 1. 전체 기본 정렬 (B2:W100)
            {"repeatCell": {"range": get_range("B2:W100"), "cell": {"userEnteredFormat": {"horizontalAlignment": "CENTER", "verticalAlignment": "MIDDLE", "textFormat": {"fontSize": 10}}}, "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,textFormat)"}},
            
            # 3. 배경색 및 글자 스타일 설정 (fields에 textFormat 추가!)
            {"repeatCell": {"range": get_range("B2:C2"), "cell": {"userEnteredFormat": {"backgroundColor": color_yellow, "textFormat": {"bold": True, "fontSize": 11}}}, "fields": "userEnteredFormat(backgroundColor,textFormat)"}},
            {"repeatCell": {"range": get_range("D2:D2"), "cell": {"userEnteredFormat": {"backgroundColor": color_soft_blue, "textFormat": {"bold": True, "fontSize": 16}}}, "fields": "userEnteredFormat(backgroundColor,textFormat)"}},
            {"repeatCell": {"range": get_range("E2:F2"), "cell": {"userEnteredFormat": {"backgroundColor": color_title_bg, "textFormat": {"bold": True, "fontSize": 16}}}, "fields": "userEnteredFormat(backgroundColor,textFormat)"}},
            
            # G, I, L, O, R 열 (backgroundColor 뒤에 ,textFormat 추가)
            {"repeatCell": {"range": get_range("G2:V2"), "cell": {"userEnteredFormat": {"textFormat": {"bold": True, "fontSize": 13}}}, "fields": "userEnteredFormat(backgroundColor,textFormat)"}},
            {"repeatCell": {"range": get_range("G2:G2"), "cell": {"userEnteredFormat": {"backgroundColor": color_light_yellow, "textFormat": {"bold": True, "fontSize": 14}}}, "fields": "userEnteredFormat(backgroundColor,textFormat)"}},
            {"repeatCell": {"range": get_range("I2:I2"), "cell": {"userEnteredFormat": {"backgroundColor": color_light_yellow, "textFormat": {"bold": True, "fontSize": 14}}}, "fields": "userEnteredFormat(backgroundColor,textFormat)"}},
            {"repeatCell": {"range": get_range("L2:L2"), "cell": {"userEnteredFormat": {"backgroundColor": color_light_yellow, "textFormat": {"bold": True, "fontSize": 14}}}, "fields": "userEnteredFormat(backgroundColor,textFormat)"}},
            {"repeatCell": {"range": get_range("O2:O2"), "cell": {"userEnteredFormat": {"backgroundColor": color_pink, "textFormat": {"bold": True, "fontSize": 14}}}, "fields": "userEnteredFormat(backgroundColor,textFormat)"}},
            {"repeatCell": {"range": get_range("R2:R2"), "cell": {"userEnteredFormat": {"backgroundColor": color_pink, "textFormat": {"bold": True, "fontSize": 14}}}, "fields": "userEnteredFormat(backgroundColor,textFormat)"}},
            
            # 4. 4행 구역별 색상 및 글자 스타일 (fields에 textFormat 추가!)
            {"repeatCell": {"range": get_range("D4:G4"), "cell": {"userEnteredFormat": {"backgroundColor": color_mint, "textFormat": {"bold": True, "fontSize": 13}}}, "fields": "userEnteredFormat(backgroundColor,textFormat)"}},
            {"repeatCell": {"range": get_range("H4:J4"), "cell": {"userEnteredFormat": {"backgroundColor": color_orange, "textFormat": {"bold": True, "fontSize": 13}}}, "fields": "userEnteredFormat(backgroundColor,textFormat)"}},
            {"repeatCell": {"range": get_range("L4:P4"), "cell": {"userEnteredFormat": {"backgroundColor": color_sky, "textFormat": {"bold": True, "fontSize": 13}}}, "fields": "userEnteredFormat(backgroundColor,textFormat)"}},
            {"repeatCell": {"range": get_range("R4:S4"), "cell": {"userEnteredFormat": {"backgroundColor": color_light_green, "textFormat": {"bold": True, "fontSize": 13}}}, "fields": "userEnteredFormat(backgroundColor,textFormat)"}},
            {"repeatCell": {"range": get_range("U4:V4"), "cell": {"userEnteredFormat": {"backgroundColor": color_lavender, "textFormat": {"bold": True, "fontSize": 13}}}, "fields": "userEnteredFormat(backgroundColor,textFormat)"}},
            
            # 5. 5행 헤더
            {"repeatCell": {"range": get_range("B5:V5"), "cell": {"userEnteredFormat": {"backgroundColor": color_header_bg, "textFormat": {"fontSize": 11, "bold": True}}}, "fields": "userEnteredFormat(backgroundColor,textFormat)"}},

            # 5. 6행부터 데이터값 설정
            {"repeatCell": {"range": get_range("B6:V36"), "cell": {"userEnteredFormat": {"textFormat": {"fontSize": 10}}}, "fields": "userEnteredFormat(backgroundColor,textFormat)"}},
        ]

        total_sum_ranges = ["E2", "H2", "J2", "M2", "P2", "T2"]
        for rng in total_sum_ranges:
            design_requests.append({
                "repeatCell": {
                    "range": get_range(rng),
                    "cell": {"userEnteredFormat": {"numberFormat": {"type": "NUMBER", "pattern": "#,##0"}, "horizontalAlignment": "CENTER"}},
                    "fields": "userEnteredFormat(numberFormat,horizontalAlignment)"
                }
            })

        data_detail_ranges = ["D6:P50", "U6:V50"]
        for rng in data_detail_ranges:
            design_requests.append({
                "repeatCell": {
                    "range": get_range(rng),
                    "cell": {"userEnteredFormat": {"numberFormat": {"type": "NUMBER", "pattern": "#,##0"}, "horizontalAlignment": "RIGHT"}},
                    "fields": "userEnteredFormat(numberFormat,horizontalAlignment)"
                }
            })

        sh.batch_update({"requests": design_requests})

        print(f"✅ {current_month_name} 시트 생성 및 최적화 설정 완료!")
        return worksheet

@web.route('/api/booking/save-google-sheet', methods=['POST'])
def save_booking_endpoint():
    data = request.json
    target_date = data.get('date') # 예: "2026-05-12"
    
    totals = data.get('totals') or {}
    combined = totals.get('combined', {})
    team = totals.get('team', {})
    supply = totals.get('supply', {})
    usage = totals.get('usageStats', {})
    print(f"📅 {target_date} 데이터 수신: {totals}")
    
    try:
        # 1. 시트 가져오기
        worksheet = save_to_google_sheet(data)
        
        day_val = int(target_date.split('-')[2])
        target_row = day_val + 5 
        days_kr = ['월', '화', '수', '목', '금', '토', '일']
        day_of_week = days_kr[datetime.strptime(target_date, '%Y-%m-%d').weekday()]

        pass_count = usage.get('passCount', 0)
        coupon_count = usage.get('couponCount', 0)
        totalUsageAmount = usage.get('totalUsageAmount', 0)
        usage_counts_text = f"{pass_count}:{coupon_count} / {int(totalUsageAmount):,}"

        # 업데이트할 값 리스트 (B열부터 W열까지)
        row_values = [
            day_val,                                         # B: 일자
            day_of_week,                                     # C: 요일
            int(combined.get('deposit', 0)),                 # D: 예약금
            int(team.get('card', 0)),                        # E: 카드(게임)
            int(team.get('cash', 0)),                        # F: 현금(게임)
            int(team.get('transfer', 0)),                    # G: 계좌(게임)
            int(supply.get('card', 0)),                      # H: 카드(기타)
            int(supply.get('cash', 0)),                      # I: 현금(기타)
            int(supply.get('transfer', 0)),                  # J: 계좌(기타)
            "",                                              # K: (공백)
            int(combined.get('card', 0)) + int(combined.get('deposit', 0)), # L: 예약금+카드
            int(combined.get('cash', 0)),                    # M: 현금
            int(combined.get('transfer', 0)),                # N: 계좌
            usage_counts_text,                                     # O: 사용다회/쿠폰 (텍스트)
            int(combined.get('total', 0)),                   # P: 총매출
            "",                                              # Q: (공백)
            int(combined.get('gameCount', 0)),               # R: 팀수
            int(combined.get('totalUsers', 0)),              # S: 인원수
            "",                                              # T: (공백)
            int(combined.get('partyRoom', 0)),               # U: 파티룸
            int(combined.get('couponTotal', 0)),             # V: 다회권
            ""                                               # W: 단체
        ]

        # 5. 특정 범위 한 번에 업데이트 (속도 향상을 위해 범위를 지정해서 쏩니다)
        range_label = f"B{target_row}:W{target_row}"
        worksheet.update(range_name=range_label, values=[row_values])

        return jsonify({'status': 'success', 'message': f'{target_date} 데이터 저장 완료!'})

    except Exception as e:
        import traceback
        print(f"저장 중 오류 발생: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500



def get_next_log_filepath(log_dir):
    """
    📅 오늘 날짜를 구한 뒤, 폴더 내 기존 파일들을 검사하여
    서버 재구동 시 자동으로 순번(_1, _2...)을 붙여주는 마법의 함수
    """
    if not os.path.exists(log_dir):
        os.makedirs(log_dir)
        
    # 1. 오늘 날짜 포맷팅 (예: 2026-05-25)
    today_str = datetime.now().strftime('%Y-%m-%d')
    
    # 2. 오늘 날짜로 시작하는 파일 번호 매기기 루프
    sequence = 1
    while True:
        # 최종 파일명 조립: 2026-05-25_1.log, 2026-05-25_2.log ...
        filename = f"{today_str}_{sequence}.log"
        full_path = os.path.join(log_dir, filename)
        
        # 만약 이 파일명이 존재하지 않는다면? ➡️ 서버가 방금 켜진 것이므로 이 번호 낚채기!
        if not os.path.exists(full_path):
            return full_path
        
        # 이미 파일이 존재하면 순번을 하나 올려서 다음 번호 검사
        sequence += 1

# 🚀 3. 서버가 켜지는 이 순간, 새롭게 배정받을 로그 파일 경로 확정!
CURRENT_LOG_PATH = get_next_log_filepath(LOG_DIR)

# 4. 로그 기록 포맷(형식) 설정
log_formatter = logging.Formatter(
    '[%(asctime)s] %(levelname)s in %(module)s: %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

# 5. 파일 저장 핸들러 지정 (하루 종일 이 파일 하나에만 순수하게 기록 누적)
# 서버가 켜질 때 이미 새 파일명을 확보했으므로 용량 제한 없이 편하게 쌓으면 됩니다.
file_handler = logging.FileHandler(CURRENT_LOG_PATH, encoding='utf-8')
file_handler.setFormatter(log_formatter)
file_handler.setLevel(logging.INFO)

# 6. 콘솔(VS Code 터미널) 출력 핸들러 세팅
console_handler = logging.StreamHandler()
console_handler.setFormatter(log_formatter)
console_handler.setLevel(logging.INFO)

# 7. Flask 기본 로거 및 웹 서버(Werkzeug) 로거에 결합
web.logger.addHandler(file_handler)
web.logger.addHandler(console_handler)
web.logger.setLevel(logging.INFO)

logging.getLogger('werkzeug').addHandler(file_handler)
logging.getLogger('werkzeug').setLevel(logging.INFO)

print(f"💾 [시스템 고도화] 이번 세션 로그가 생성되었습니다: {CURRENT_LOG_PATH}")


def get_today_log_file():
    today = datetime.now().strftime("%Y_%m_%d")

    if not os.path.exists(SERVER_LOG_DIR):
        print(f"❌ 로그 폴더 없음: {SERVER_LOG_DIR}")
        return None

    candidates = []
    for f in os.listdir(SERVER_LOG_DIR):
        if f.startswith(today) and f.endswith("_log.txt"):
            candidates.append(f)

    if not candidates:
        return None

    # 가장 최근 파일 선택
    candidates.sort(reverse=True)
    selected = candidates[0]

    print(f"📌 선택된 로그 파일: {selected}")   # ✅ 추가
    return os.path.join(SERVER_LOG_DIR, candidates[0])

def open_log_file(path):
    print(f"📄 로그 파일 열기: {path}")
    return open(path, "r", encoding="utf-8", errors="ignore")

   
    
# ========================================================
# 콘솔 UTF-8
# ========================================================
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except:
    pass

# ========================================================
# PAD 매핑
# ========================================================
PAD_MAP = {
    0: "C1",
    1: "C2",
    2: "B1",
    3: "B2",
}

# ========================================================
# map_id → size / level
# ========================================================
MAP_INFO = {
    236: {"size": "소형", "level": "Basic"},
    233: {"size": "소형", "level": "Easy"},
    234: {"size": "소형", "level": "Normal"},
    237: {"size": "소형", "level": "Hard"},
    254: {"size": "소형", "level": "Challenger"},
    255: {"size": "소형", "level": "Space"},
    262: {"size": "소형", "level": "Summer"},
    267: {"size": "소형", "level": "Kids"},
    270: {"size": "소형", "level": "Santa"},
    231: {"size": "중형", "level": "Basic"},
    215: {"size": "중형", "level": "Easy"},
    209: {"size": "중형", "level": "Normal"},
    214: {"size": "중형", "level": "Hard"},
    253: {"size": "중형", "level": "Challenger"},
    252: {"size": "중형", "level": "Space"},
    261: {"size": "중형", "level": "Summer"},
    266: {"size": "중형", "level": "Kids"},
    269: {"size": "중형", "level": "Santa"},
    221: {"size": "대형", "level": "Basic"},
    213: {"size": "대형", "level": "Easy"},
    210: {"size": "대형", "level": "Normal"},
    212: {"size": "대형", "level": "Hard"},
    256: {"size": "대형", "level": "Challenger"},
    217: {"size": "대형", "level": "Space"},
    250: {"size": "대형", "level": "Summer"},
    222: {"size": "대형", "level": "Kids"},
    265: {"size": "대형", "level": "Santa"},
}

LEVEL_KR = {
    "Basic": "베이직",
    "Easy": "이지",
    "Normal": "노멀",
    "Hard": "하드",
    "Challenger": "챌린저",
    "Kids": "유아",
    "Summer": "여름",
    "Space": "우주",
    "Santa": "산타",
    "all": "전체"
}

# ========================================================
# 로그 통계
# ========================================================
log_stats = {
    "db_insert_success": 0,
    "db_insert_duplicate": 0,
}

# ========================================================
# 파싱 통계
# ========================================================
parse_stats = {
    "total": 0,
    "start": 0,
    "stop": 0,
    "rank": 0,
    "failed": 0,
}

# ========================================================
# DB 초기화
# ========================================================
def init_db():
    conn = get_db_connection()
    cur = conn.cursor()

    # WAL keeps readers responsive while the log monitor records game results.
    cur.execute("PRAGMA journal_mode = WAL")
    cur.execute("PRAGMA synchronous = NORMAL")

    cur.execute('''CREATE TABLE IF NOT EXISTS supply_history (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        target_date TEXT,
        time TEXT,
        item TEXT,
        quantity INTEGER,
        etc_text TEXT,
        card_amount INTEGER,
        cash_amount INTEGER,
        transfer_amount INTEGER,
        total_amount INTEGER,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')

    cur.execute('''CREATE TABLE IF NOT EXISTS settlement_daily_meta (
        target_date TEXT PRIMARY KEY,
        cash_expense INTEGER DEFAULT 0,
        no_show_count INTEGER DEFAULT 0,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')

    cur.execute("PRAGMA table_info(settlement_daily_meta)")
    settlement_meta_columns = {row[1] for row in cur.fetchall()}
    if 'cash_expense' not in settlement_meta_columns:
        cur.execute("ALTER TABLE settlement_daily_meta ADD COLUMN cash_expense INTEGER DEFAULT 0")
    if 'no_show_count' not in settlement_meta_columns:
        cur.execute("ALTER TABLE settlement_daily_meta ADD COLUMN no_show_count INTEGER DEFAULT 0")

    cur.execute("PRAGMA table_info(supply_history)")
    supply_columns = {row[1] for row in cur.fetchall()}
    if 'target_date' not in supply_columns:
        cur.execute("ALTER TABLE supply_history ADD COLUMN target_date TEXT")
    if 'quantity' not in supply_columns:
        cur.execute("ALTER TABLE supply_history ADD COLUMN quantity INTEGER")
    if 'etc_text' not in supply_columns:
        cur.execute("ALTER TABLE supply_history ADD COLUMN etc_text TEXT")
    cur.execute("""
        UPDATE supply_history
           SET target_date = COALESCE(NULLIF(target_date, ''), NULLIF(substr(created_at, 1, 10), ''), date('now', 'localtime'))
         WHERE target_date IS NULL OR target_date = ''
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS game_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            time TEXT,
            pad_id INTEGER,
            map_id INTEGER,
            size TEXT,
            level TEXT,
            team TEXT,
            score INTEGER,
            UNIQUE(time, pad_id, score)
        )
    """)

    cur.execute('''CREATE TABLE IF NOT EXISTS walkins
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  name TEXT NOT NULL,
                  team TEXT,
                  level TEXT,
                  people INTEGER,
                  room_size TEXT,
                  room_fast INTEGER DEFAULT 0,
                  phone TEXT,
                  is_agreed INTEGER,
                  reg_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                  status TEXT DEFAULT 'waiting')''')

    cur.execute("PRAGMA table_info(walkins)")
    walkin_columns = {row[1] for row in cur.fetchall()}
    if 'room_fast' not in walkin_columns:
        cur.execute("ALTER TABLE walkins ADD COLUMN room_fast INTEGER DEFAULT 0")
    if 'adult_count' not in walkin_columns:
        cur.execute("ALTER TABLE walkins ADD COLUMN adult_count INTEGER DEFAULT 0")
    if 'child_count' not in walkin_columns:
        cur.execute("ALTER TABLE walkins ADD COLUMN child_count INTEGER DEFAULT 0")
    if 'initial_level' not in walkin_columns:
        cur.execute("ALTER TABLE walkins ADD COLUMN initial_level TEXT")
    if 'initial_room_size' not in walkin_columns:
        cur.execute("ALTER TABLE walkins ADD COLUMN initial_room_size TEXT")
    if 'initial_room_fast' not in walkin_columns:
        cur.execute("ALTER TABLE walkins ADD COLUMN initial_room_fast INTEGER")

    # The staff list is an intake record.  Keep its initial choice separate
    # from a game card's later operational edits.
    intake_level_col = 'level' if 'level' in walkin_columns else 'diff'
    cur.execute(
        f"UPDATE walkins SET initial_level={intake_level_col} "
        "WHERE initial_level IS NULL OR initial_level=''"
    )
    cur.execute(
        "UPDATE walkins SET initial_room_size=room_size "
        "WHERE initial_room_size IS NULL OR initial_room_size=''"
    )
    cur.execute(
        "UPDATE walkins SET initial_room_fast=room_fast "
        "WHERE initial_room_fast IS NULL"
    )

    cur.execute("""
        CREATE TABLE IF NOT EXISTS account_deposit (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            deposit_date DATE NOT NULL,
            deposit_time TIME NOT NULL,
            name TEXT NOT NULL,
            amount INTEGER NOT NULL
        );
        """)    

    cur.execute("PRAGMA table_info(bookings)")
    booking_columns = {row[1] for row in cur.fetchall()}
    if not booking_columns:
        cur.execute('''CREATE TABLE bookings
                     (id INTEGER PRIMARY KEY AUTOINCREMENT,
                      booking_date TEXT NOT NULL,
                      time_key TEXT NOT NULL,
                      room TEXT NOT NULL,
                      name TEXT,
                      phone TEXT,
                      team TEXT,
                      level TEXT,
                      people TEXT,
                      order_no INTEGER DEFAULT 0,
                      paid INTEGER DEFAULT 0,
                      completed INTEGER DEFAULT 0,
                      payment_data TEXT,
                      created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
    else:
        if 'booking_date' not in booking_columns:
            cur.execute("ALTER TABLE bookings ADD COLUMN booking_date TEXT")
        if 'order_no' not in booking_columns:
            cur.execute("ALTER TABLE bookings ADD COLUMN order_no INTEGER DEFAULT 0")
        if 'paid' not in booking_columns:
            cur.execute("ALTER TABLE bookings ADD COLUMN paid INTEGER DEFAULT 0")
        if 'completed' not in booking_columns:
            cur.execute("ALTER TABLE bookings ADD COLUMN completed INTEGER DEFAULT 0")
        if 'payment_data' not in booking_columns:
            cur.execute("ALTER TABLE bookings ADD COLUMN payment_data TEXT")
        if 'created_at' not in booking_columns:
            cur.execute("ALTER TABLE bookings ADD COLUMN created_at TEXT")

        cur.execute("UPDATE bookings SET created_at = COALESCE(NULLIF(created_at, ''), datetime('now', 'localtime')) WHERE created_at IS NULL OR created_at = ''")
        cur.execute("UPDATE bookings SET booking_date = COALESCE(NULLIF(booking_date, ''), NULLIF(substr(created_at, 1, 10), ''), date('now', 'localtime')) WHERE booking_date IS NULL OR booking_date = ''")

        cur.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='bookings'")
        tbl_sql = (cur.fetchone() or ('',))[0]
        if 'UNIQUE(time_key, room)' in tbl_sql:
            cur.execute('''CREATE TABLE bookings_new
                         (id INTEGER PRIMARY KEY AUTOINCREMENT,
                          booking_date TEXT NOT NULL,
                          time_key TEXT NOT NULL,
                          room TEXT NOT NULL,
                          name TEXT,
                          phone TEXT,
                          team TEXT,
                          level TEXT,
                          people TEXT,
                          order_no INTEGER DEFAULT 0,
                          paid INTEGER DEFAULT 0,
                          completed INTEGER DEFAULT 0,
                          payment_data TEXT,
                          created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
            cur.execute('''INSERT INTO bookings_new
                        (id, booking_date, time_key, room, name, phone, team, level, people, order_no, paid, completed, payment_data, created_at)
                         SELECT id,
                                COALESCE(NULLIF(booking_date, ''), NULLIF(substr(created_at, 1, 10), ''), date('now', 'localtime')),
                                time_key, room, name, phone, team, level, people,
                                COALESCE(order_no, 0), COALESCE(paid, 0), COALESCE(completed, 0), payment_data,
                                COALESCE(NULLIF(created_at, ''), datetime('now', 'localtime'))
                         FROM bookings''')
            cur.execute('DROP TABLE bookings')
            cur.execute('ALTER TABLE bookings_new RENAME TO bookings')

    cur.execute('''CREATE TABLE IF NOT EXISTS day_type_overrides
                 (target_date TEXT PRIMARY KEY,
                  day_type TEXT NOT NULL,
                  updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')

    cur.execute('''CREATE TABLE IF NOT EXISTS queue_items
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  room TEXT NOT NULL,
                  order_no INTEGER DEFAULT 0,
                  name TEXT,
                  team TEXT,
                  level TEXT,
                  bid INTEGER DEFAULT 0,
                  party_room INTEGER DEFAULT 0,
                  people TEXT,
                  created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')


    cur.execute('''CREATE TABLE IF NOT EXISTS naver_mail_cache (
                booking_id TEXT PRIMARY KEY,
                masked_name TEXT NOT NULL,
                use_date TEXT NOT NULL,
                use_time_key TEXT NOT NULL,
                room_name TEXT NOT NULL,
                status TEXT DEFAULT 'INIT',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
        

    # Logical rooms remain stable even when each room later uses a separate PC.
    cur.execute('''CREATE TABLE IF NOT EXISTS rooms (
                room_id TEXT PRIMARY KEY,
                display_name TEXT NOT NULL,
                pad_id INTEGER,
                agent_mode TEXT NOT NULL DEFAULT 'shared_manager',
                enabled INTEGER NOT NULL DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
    cur.executemany(
        '''INSERT OR IGNORE INTO rooms (room_id, display_name, pad_id)
           VALUES (?, ?, ?)''',
        [(room_name, room_name, pad_id) for pad_id, room_name in PAD_MAP.items()],
    )

    cur.execute('''CREATE TABLE IF NOT EXISTS room_agents (
                agent_id TEXT PRIMARY KEY,
                room_id TEXT NOT NULL,
                agent_name TEXT NOT NULL DEFAULT '',
                host_name TEXT NOT NULL DEFAULT '',
                agent_version TEXT NOT NULL DEFAULT '',
                connection_mode TEXT NOT NULL DEFAULT 'mqtt',
                status TEXT NOT NULL DEFAULT 'offline',
                last_seen_at TEXT,
                capabilities_json TEXT NOT NULL DEFAULT '[]',
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(room_id) REFERENCES rooms(room_id))''')

    cur.execute('''CREATE TABLE IF NOT EXISTS command_queue (
                command_id TEXT PRIMARY KEY,
                room_id TEXT NOT NULL,
                command_type TEXT NOT NULL,
                payload_json TEXT NOT NULL DEFAULT '{}',
                status TEXT NOT NULL DEFAULT 'pending',
                requested_by TEXT NOT NULL DEFAULT '',
                requested_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                expires_at TIMESTAMP,
                claimed_by TEXT,
                claimed_at TIMESTAMP,
                completed_at TIMESTAMP,
                result_json TEXT NOT NULL DEFAULT '{}',
                error_message TEXT NOT NULL DEFAULT '',
                FOREIGN KEY(room_id) REFERENCES rooms(room_id))''')

    cur.execute('''CREATE TABLE IF NOT EXISTS command_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                command_id TEXT NOT NULL,
                room_id TEXT NOT NULL,
                event_type TEXT NOT NULL,
                actor_id TEXT NOT NULL DEFAULT '',
                detail_json TEXT NOT NULL DEFAULT '{}',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')

    cur.execute('''CREATE TABLE IF NOT EXISTS naver_reservations (
                booking_id TEXT PRIMARY KEY,
                booking_status TEXT NOT NULL,
                use_date TEXT NOT NULL,
                use_time_key TEXT NOT NULL,
                room_name TEXT NOT NULL DEFAULT '',
                product_name TEXT NOT NULL DEFAULT '',
                customer_name TEXT NOT NULL DEFAULT '',
                team_name TEXT NOT NULL DEFAULT '',
                difficulty TEXT NOT NULL DEFAULT '',
                phone TEXT NOT NULL DEFAULT '',
                people_count INTEGER,
                booking_fingerprint TEXT NOT NULL DEFAULT '',
                first_seen_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                cancelled_at TIMESTAMP)''')

    cur.execute("PRAGMA table_info(naver_reservations)")
    naver_reservation_columns = {row[1] for row in cur.fetchall()}
    for column_name in ("team_name", "difficulty", "phone"):
        if column_name not in naver_reservation_columns:
            cur.execute(f"ALTER TABLE naver_reservations ADD COLUMN {column_name} TEXT NOT NULL DEFAULT ''")

    cur.execute('''CREATE TABLE IF NOT EXISTS naver_stock_rules (
                rule_id INTEGER PRIMARY KEY AUTOINCREMENT,
                business_item_id TEXT NOT NULL,
                use_date TEXT NOT NULL,
                use_time_key TEXT NOT NULL,
                room_id TEXT NOT NULL DEFAULT '',
                blocked INTEGER NOT NULL DEFAULT 1,
                reason TEXT NOT NULL DEFAULT '',
                updated_by TEXT NOT NULL DEFAULT '',
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(business_item_id, use_date, use_time_key))''')

    cur.execute('''CREATE TABLE IF NOT EXISTS naver_stock_sync_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                rule_id INTEGER,
                action TEXT NOT NULL,
                status TEXT NOT NULL,
                response_summary TEXT NOT NULL DEFAULT '',
                synced_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')

    # Keeps only a booking identifier to prevent duplicate same-day cancellation counts.
    cur.execute('''CREATE TABLE IF NOT EXISTS naver_cancellation_events (
                booking_id TEXT PRIMARY KEY,
                use_date TEXT NOT NULL,
                counted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')

    cur.execute('''CREATE TABLE IF NOT EXISTS naver_booking_card_links (
                booking_id TEXT PRIMARY KEY,
                booking_row_id INTEGER,
                handling_mode TEXT NOT NULL DEFAULT 'standard',
                card_state TEXT NOT NULL DEFAULT 'active',
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(booking_row_id) REFERENCES bookings(id))''')

    cur.execute('''CREATE TABLE IF NOT EXISTS db_maintenance_runs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                job_name TEXT NOT NULL,
                status TEXT NOT NULL,
                detail TEXT NOT NULL DEFAULT '',
                started_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                completed_at TIMESTAMP)''')

    # Existing-table indexes: additive only, safe for the current schema.
    cur.execute('CREATE INDEX IF NOT EXISTS idx_game_records_time_pad ON game_records(time, pad_id)')
    cur.execute('CREATE INDEX IF NOT EXISTS idx_bookings_date_room_time ON bookings(booking_date, room, time_key)')
    cur.execute('CREATE INDEX IF NOT EXISTS idx_queue_items_room_order ON queue_items(room, order_no)')
    # Old operating databases use visit_date; newer empty databases use reg_time.
    if 'visit_date' in walkin_columns:
        cur.execute('CREATE INDEX IF NOT EXISTS idx_walkins_date_status ON walkins(visit_date, status)')
    elif 'reg_time' in walkin_columns:
        cur.execute('CREATE INDEX IF NOT EXISTS idx_walkins_time_status ON walkins(reg_time, status)')
    cur.execute('CREATE INDEX IF NOT EXISTS idx_naver_mail_cache_today ON naver_mail_cache(use_date, status, use_time_key)')
    cur.execute('CREATE INDEX IF NOT EXISTS idx_supply_history_date ON supply_history(target_date)')
    cur.execute('CREATE INDEX IF NOT EXISTS idx_room_agents_room_status ON room_agents(room_id, status)')
    cur.execute('CREATE INDEX IF NOT EXISTS idx_command_queue_room_status ON command_queue(room_id, status, requested_at)')
    cur.execute('CREATE INDEX IF NOT EXISTS idx_command_history_command ON command_history(command_id, created_at)')
    cur.execute('CREATE INDEX IF NOT EXISTS idx_naver_reservations_date_status ON naver_reservations(use_date, booking_status, use_time_key)')
    cur.execute('CREATE INDEX IF NOT EXISTS idx_naver_booking_card_links_row ON naver_booking_card_links(booking_row_id)')
    cur.execute('CREATE INDEX IF NOT EXISTS idx_naver_stock_rules_date ON naver_stock_rules(use_date, use_time_key)')

    cur.execute("PRAGMA table_info(queue_items)")
    queue_cols = [r[1] for r in cur.fetchall()]
    if 'bid' not in queue_cols:
        cur.execute("ALTER TABLE queue_items ADD COLUMN bid INTEGER DEFAULT 0")
    if 'party_room' not in queue_cols:
        cur.execute("ALTER TABLE queue_items ADD COLUMN party_room INTEGER DEFAULT 0")
    if 'order_no' not in queue_cols:
        cur.execute("ALTER TABLE queue_items ADD COLUMN order_no INTEGER DEFAULT 0")
        cur.execute("UPDATE queue_items SET order_no = id WHERE COALESCE(order_no, 0) = 0")

    conn.commit()
    conn.close()


# ========================================================
# Web - Supply History (기타판매 결제리스트)
# ========================================================
from flask import abort

@web.route('/api/supply_history/save', methods=['POST'])
def save_supply_history():
    data = request.json
    if not isinstance(data, list):
        return jsonify({'status': 'error', 'message': '리스트 형식이 아님'}), 400
    target_date = _get_request_date(default_today=True)
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    cur.execute('DELETE FROM supply_history WHERE target_date = ?', (target_date,))
    for entry in data:
        cur.execute('''INSERT INTO supply_history (target_date, time, item, quantity, etc_text, card_amount, cash_amount, transfer_amount, total_amount)
                      VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''', (
            target_date,
            entry.get('time'),
            entry.get('item'),
            int(entry.get('quantity')) if entry.get('quantity') is not None and str(entry.get('quantity')).strip() != '' else None,
            entry.get('etcText') or None,
            int(entry.get('cardAmount', 0)),
            int(entry.get('cashAmount', 0)),
            int(entry.get('transferAmount', 0)),
            int(entry.get('totalAmount', 0))
        ))
    conn.commit()
    conn.close()
    return jsonify({'status': 'success'})

@web.route('/api/supply_history/<int:sid>', methods=['PUT'])
def update_supply_item(sid):
    data = request.json
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    try:
        # 필드 정리 (JS에서 보낸 키값과 매칭)
        time = data.get('time', '')
        item = data.get('item', '항목')
        etc_text = data.get('etc_text', '')

        qty_val = data.get('quantity')
        if qty_val is None or str(qty_val).strip() == "" or str(qty_val) == "NaN":
            parsed_qty = None
        else:
            try:
                parsed_qty = int(qty_val)
            except (ValueError, TypeError):
                parsed_qty = None

        card = int(data.get('card_amount') or 0)
        cash = int(data.get('cash_amount') or 0)
        transfer = int(data.get('transfer_amount') or 0)
        total = card + cash + transfer

        cur.execute('''
            UPDATE supply_history 
            SET time=?, item=?, etc_text=?, quantity=?, 
                card_amount=?, cash_amount=?, transfer_amount=?, total_amount=?
            WHERE id=?
        ''', (
            time,           # 1
            item,           # 2
            etc_text,       # 3
            parsed_qty,     # 4
            card,           # 5
            cash,           # 6
            transfer,       # 7
            total,          # 8
            sid             # 9 (WHERE id=?)
        ))
        
        conn.commit()
        return jsonify({'status': 'success'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        conn.close()

# 2. 개별 항목 삭제
@web.route('/api/supply_history/<int:sid>', methods=['DELETE'])
def delete_supply_item(sid):
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    try:
        cur.execute('DELETE FROM supply_history WHERE id=?', (sid,))
        conn.commit()
        return jsonify({'status': 'success'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        conn.close()

# 3. 신규 항목 추가 (기존 POST 방식 활용 또는 신규 생성)
@web.route('/api/supply_history', methods=['POST'])
def create_supply_item():
    data = request.json
    target_date = data.get('target_date') or _get_request_date(default_today=True)
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    try:
        cur.execute('''
            INSERT INTO supply_history (target_date, time, item, etc_text, quantity, card_amount, cash_amount, transfer_amount, total_amount)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (target_date, '', '항목', '', None, 0, 0, 0, 0))
        new_id = cur.lastrowid
        conn.commit()
        return jsonify({'status': 'success', 'id': new_id})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        conn.close()


@web.route('/api/supply_history/list', methods=['GET'])
def get_supply_history():
    target_date = _get_request_date(default_today=True)
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute('SELECT * FROM supply_history WHERE target_date = ? ORDER BY id ASC', (target_date,))
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return jsonify(rows)


def ensure_queue_items_order_column(conn):
    cur = conn.cursor()
    cur.execute("PRAGMA table_info(queue_items)")
    cols = [r[1] for r in cur.fetchall()]
    if 'order_no' not in cols:
        cur.execute("ALTER TABLE queue_items ADD COLUMN order_no INTEGER DEFAULT 0")
        cur.execute("UPDATE queue_items SET order_no = id WHERE COALESCE(order_no, 0) = 0")
        conn.commit()


def _today_date_str():
    return datetime.now().strftime('%Y-%m-%d')


def _normalize_date_str(date_str):
    value = str(date_str or '').strip()
    try:
        return datetime.strptime(value, '%Y-%m-%d').strftime('%Y-%m-%d')
    except Exception:
        return None


def _get_request_date(default_today=True):
    raw = request.args.get('date') or request.values.get('date')
    normalized = _normalize_date_str(raw)
    if normalized:
        return normalized
    return _today_date_str() if default_today else None


def _parse_payment_data_safe(raw):
    if raw is None:
        return {}
    if isinstance(raw, dict):
        return raw
    if not isinstance(raw, str):
        return {}
    try:
        parsed = json.loads(raw)
        return parsed if isinstance(parsed, dict) else {}
    except Exception:
        pass
    try:
        normalized = (
            raw.replace('None', 'null')
               .replace('True', 'true')
               .replace('False', 'false')
               .replace("'", '"')
        )
        parsed = json.loads(normalized)
        return parsed if isinstance(parsed, dict) else {}
    except Exception:
        return {}



# ========================================================
# 로그 전체 1회 스캔
# ========================================================
def scan_existing_log_and_save(path):
    print("🔁 기존 로그 전체 스캔 (DB 미저장 보정)")

    # 상태 초기화
    for i in range(PAD_COUNT):
        pad_status[i] = "off"
        pad_data[i] = init_pad_data(i)

    with open(path, "rb") as f:
        for raw in f:
            line = safe_decode(raw).strip()
            if not line:
                continue

            parse_stats["total"] += 1
            parse_log_line(line)

    print(
        f"🔁 스캔 완료 | "
        f"START:{parse_stats['start']} "
        f"STOP:{parse_stats['stop']} "
        f"DUP:{log_stats['db_insert_duplicate']}"
    )
            
def restore_pad_state_from_log(path):
    print("🔁 기존 로그 스캔 (재시작 복구)")

    runtime = {i: None for i in range(PAD_COUNT)}
    last_stop = None

    with open(path, "rb") as f:
        for raw in f:
            try:
                line = raw.decode("utf-8")
            except:
                line = raw.decode("cp949", errors="ignore")

            m = GAME_START_PATTERN.search(line)
            if m:
                pad_id = int(m.group(1))
                map_id = int(m.group(2))
                t = TIME_PATTERN.search(line)
                runtime[pad_id] = {
                    "map_id": map_id,
                    "time": normalize_time(t.group(1)) if t else None
                }
                continue

            m = GAME_STOP_PATTERN.search(line)
            if m:
                last_stop = int(m.group(1))
                runtime[last_stop] = None
                continue


    for pid, info in runtime.items():
        if info:
            pad_data[pid] = init_pad_data(pid)
            pad_data[pid]["map_id"] = info["map_id"]
            pad_data[pid]["time"] = info["time"]
            if info["map_id"] in MAP_INFO:
                pad_data[pid].update(MAP_INFO[info["map_id"]])
            pad_status[pid] = "playing"
            print(f"♻ PAD 복구 | {PAD_MAP[pid]} | playing")
            print(
                f"▶ GAME START | {pad_data[pid]['time']} | {PAD_MAP[pid]} | "
                f"{pad_data[pid]['size']} {pad_data[pid]['level']}"
            )
        else:
            pad_status[pid] = "off"


def parse_log_line(line: str):
    """
    로그 1줄을 처리한다.
    scan_existing_log_and_save / log_monitor 에서 공통 사용
    """
    global pad_status, pad_data

    matched = False

    # =========================
    # ▶ GAME START
    # =========================
    m = GAME_START_PATTERN.search(line)
    if m:
        matched = True
        parse_stats["start"] += 1

        pad_id = int(m.group(1))
        map_id = int(m.group(2))

        pad_data[pad_id] = init_pad_data(pad_id)
        pad_data[pad_id]["map_id"] = map_id

        t = TIME_PATTERN.search(line)
        if t:
            pad_data[pad_id]["time"] = normalize_time(t.group(1))
        else:
            pad_data[pad_id]["time"] = None

        if map_id in MAP_INFO:
            pad_data[pad_id].update(MAP_INFO[map_id])

        pad_status[pad_id] = "playing"

        print(
            f"▶ GAME START | {pad_data[pad_id]['time']} | {PAD_MAP[pad_id]} | "
            f"{pad_data[pad_id]['size']} {pad_data[pad_id]['level']}"
        )
        # 실시간 대시보드 갱신
        socketio.emit('room_or_queue_changed')
        return True

    # =========================
    # ⏹ GAME STOP
    # =========================
    m = GAME_STOP_PATTERN.search(line)
    if m:
        matched = True
        parse_stats["stop"] += 1

        pad_id = int(m.group(1))

        pad_status[pad_id] = "wait_rank"
        pad_data[pad_id]["score"] = None

        t = TIME_PATTERN.search(line)
        if t:
            pad_data[pad_id]["time"] = normalize_time(t.group(1))
        else:
            pad_data[pad_id]["time"] = None

        print(f"⏹ STOP | {pad_data[pad_id]['time']} | {PAD_MAP[pad_id]}")
        # 실시간 대시보드 갱신
        socketio.emit('room_or_queue_changed')
        return True

    # =========================
    # 🎖 FINAL RANK
    # =========================
    r = FINAL_RANK_PATTERN.search(line)
    if r:
        matched = True
        parse_stats["rank"] += 1

        team = r.group("team")
        score = int(r.group("score"))

        t = TIME_PATTERN.search(line)

        rank_time = normalize_time(t.group(1)) if t else None
        if not rank_time:
            return True

        rank_dt = parse_time(rank_time)

        waiting_pads = [
            pid for pid, s in pad_status.items()
            if s == "wait_rank" and pad_data[pid]["time"]
        ]

        if not waiting_pads:
            print("⚠ RANK 수신됐지만 대기 pad 없음")
            return True

        pad_id = min(
            waiting_pads,
            key=lambda pid: abs(
                rank_dt - parse_time(pad_data[pid]["time"])
            )
        )

        pad_data[pad_id]["team"] = team
        pad_data[pad_id]["score"] = score

        save_to_db(pad_data[pad_id])

        pad_status[pad_id] = "off"

        print(
            f"✅ SAVE | {PAD_MAP[pad_id]} | "
            f"[{team}] {score}점"
        )
        return True

    # =========================
    # 🚫 RANK SKIP
    # =========================
    if RANK_SKIP_PATTERN.search(line):
        matched = True

        waiting_pads = [
            pid for pid, s in pad_status.items()
            if s == "wait_rank"
        ]

        if waiting_pads:
            pad_id = waiting_pads[0]
            pad_status[pad_id] = "off"

            print(f"🚫 RANK SKIP | {PAD_MAP[pad_id]}")
        return True

    # =========================
    # ❌ 파싱 실패 통계
    # =========================
    if not matched and ("[QUE]" in line or "[Rank]" in line):
        parse_stats["failed"] += 1

    return False



# ========================================================
# pad 상태
# ========================================================
pad_status = {i: "off" for i in range(PAD_COUNT)}

def init_pad_data(pad_id):
    return {
        "time": None,
        "pad_id": pad_id,
        "map_id": None,
        "size": None,
        "level": None,
        "team": None,
        "score": None,
    }

pad_data = {i: init_pad_data(i) for i in range(PAD_COUNT)}

# ========================================================
# 정규식
# ========================================================
TIME_PATTERN = re.compile(r"^(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})")
GAME_START_PATTERN = re.compile(r"\[QUE\].*게임 시작.*pad_id\s*[:=]\s*(\d+).*map_id\s*[:=]\s*(\d+)")
GAME_STOP_PATTERN = re.compile(r"\[QUE\]게임 정지 요청 추가.*pad_id:\s*(\d+)")
FINAL_RANK_PATTERN = re.compile(
    r"\[Rank\]\s*>>\s*\[(?P<team>[^\]]+)\].*?"
    r"지역 랭킹:\s*\d+등\s*/\s*점수:\s*(?P<score>\d+)"
)
RANK_SKIP_PATTERN = re.compile(r"\[Rank\].*랭킹 추가 안함")


def parse_time(t):
    if not t:
        return None
    try:
        if len(t) == 16:  # YYYY-MM-DD HH:MM
            return datetime.strptime(t, "%Y-%m-%d %H:%M")
        else:  # YYYY-MM-DD HH:MM:SS
            return datetime.strptime(t, "%Y-%m-%d %H:%M:%S")
    except:
        return None



def normalize_time(t: str):
    """
    로그에서 뽑은 시간 문자열을
    YYYY-MM-DD HH:MM 로 통일
    """
    if not t:
        return None
    try:
        dt = datetime.strptime(t, "%Y-%m-%d %H:%M:%S")
    except:
        try:
            dt = datetime.strptime(t, "%Y-%m-%d %H:%M")
        except:
            return None
    return dt.strftime("%Y-%m-%d %H:%M")



# ========================================================
# DB 함수
# ========================================================
def save_to_db(data):
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    cur.execute("""
        INSERT OR IGNORE INTO game_records
        (time, pad_id, map_id, size, level, team, score)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (
        data["time"], data["pad_id"], data["map_id"],
        data["size"], data["level"], data["team"],
        data["score"]
    ))

    inserted = cur.rowcount > 0

    if not inserted:
        log_stats["db_insert_duplicate"] += 1
        print(
            f"🔁 DUPLICATE | {PAD_MAP.get(data['pad_id'])} | "
            f"{data.get('time')} | score={data.get('score')}"
        )
    else:
        log_stats["db_insert_success"] += 1
        print("🔥 DB INSERT SUCCESS → 랭킹 리프레시 emit")

    conn.commit()
    conn.close()

    # ✅ 실제 INSERT 성공한 경우에만 emit
    if inserted:
        socketio.emit("rank_refresh")

    

def load_rows():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("""
        SELECT * FROM game_records
        WHERE score IS NOT NULL
        ORDER BY score DESC
    """)
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows

def load_rows_by_date(date_str):
    """
    date_str: '2025-12-23'
    """
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    cur.execute("""
        SELECT *
        FROM game_records
        WHERE time LIKE ?
        ORDER BY time DESC
    """, (f"{date_str}%",))

    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows






# ========================================================
# 로그 디코딩 (UTF-8 / CP949 자동 판별)
# ========================================================
def safe_decode(data: bytes) -> str:
    if not data:
        return ""

    try:
        return data.decode("utf-8")
    except UnicodeDecodeError:
        try:
            return data.decode("cp949")
        except UnicodeDecodeError:
            return data.decode("utf-8", errors="replace")

# ========================================================
# 로그 offset 저장 / 복구
# ========================================================
def load_offset():
    try:
        with open(OFFSET_FILE, "r") as f:
            return int(f.read().strip())
    except:
        return 0


def save_offset(offset):
    try:
        with open(OFFSET_FILE, "w") as f:
            f.write(str(offset))
    except:
        pass

# ========================================================
# 로그 follow
# ========================================================
def follow(file, start_at_end=True):
    if start_at_end:
        file.seek(0, os.SEEK_END)
    while True:
        line = file.readline()
        if not line:
            time.sleep(0.1)
            continue
        yield line
            
# ========================================================
# 로그 감시
# ========================================================
def log_monitor():
    print("📡 로그 감시 시작 (SQLite)")

    # 파싱 통계 초기화
    for k in parse_stats:
        parse_stats[k] = 0

    for k in log_stats:
        log_stats[k] = 0

    log_file = get_today_log_file()
    #log_file = r"D:\test\logs\2026_02_14__10_11_28_log.txt"
    
    if not log_file:
        print("⚠ 로그 파일 없음")
        return

    # 🔥 1단계: 과거 로그 전부 보정
    scan_existing_log_and_save(log_file)

    # 🔥 2단계: 현재 진행 중 PAD 복구
    restore_pad_state_from_log(log_file)

    # 🔥 3단계: 실시간 follow
    f = open_log_file(log_file)
    f.seek(0, os.SEEK_END)

    last_report = time.time()

    for raw in follow(f, start_at_end=False):
        if not raw:
            continue

        line = raw.strip()
        if not line:
            continue

        parse_stats["total"] += 1
        parse_log_line(line)

        if time.time() - last_report >= 10:
            print(
                f"📊 LOG STATS | "
                f"TOTAL:{parse_stats['total']} | "
                f"START:{parse_stats['start']} | "
                f"STOP:{parse_stats['stop']} | "
                f"FAILED:{parse_stats['failed']}"
            )
            last_report = time.time()


# ========================================================
# Web - Rank
# ========================================================
@web.route("/")
def rank_page():
    rows = load_rows()
    highlight_id = request.args.get("highlight")
    
    current_range = request.args.get("range", "today")
    current_level = request.args.get("level", "all")
    current_size = request.args.get("size", "all")

    now = datetime.now()
    filtered = []

    for r in rows:
        try:
            t = datetime.strptime(r["time"], "%Y-%m-%d %H:%M")
        except:
            continue

        # 기간 필터
        if current_range == "today" and t.date() != now.date():
            continue
        if current_range == "week" and t < now - timedelta(days=7):
            continue
        if current_range == "month" and t < now - timedelta(days=30):
            continue

        # 레벨 필터
        if current_level != "all" and r["level"] != current_level:
            continue

        # 사이즈 필터
        if current_size != "all" and r["size"] != current_size:
            continue

        filtered.append(r)

    # 1. Latest 데이터 자르기 (최신순)
    if current_range in ["latest20", "latest100"]:
        filtered.sort(key=lambda x: x["time"], reverse=True)
        
        limit = 20 if current_range == "latest20" else 100
        filtered = filtered[:limit]

    # 2. 점수 기준 정렬 (공통)
    # - Latest는 잘린 데이터 내에서 점수순 정렬
    # - 나머지는 전체 데이터에서 점수순 정렬
    filtered.sort(key=lambda x: x["score"], reverse=True)

    # 순위권 밖 검색결과 표기를 위해 전체 리스트 복사
    full_sorted_list = list(filtered) if highlight_id else []

    # 3. 랭킹 산정 (동점 처리 포함)
    rank = 0
    prev_score = None
    display_rank = 0

    for r in filtered:
        rank += 1
        if r["score"] != prev_score:
            display_rank = rank
            prev_score = r["score"]
        r["display_rank"] = display_rank

    # 4. 최대 표시 개수 제한 (Latest가 아닌 경우에만 1000개 제한)
    if current_range not in ["latest20", "latest100"]:
        is_highlight_in_top_1000 = False
        if highlight_id:
            for i, r in enumerate(filtered):
                if i >= 1000: break
                if str(r['id']) == highlight_id:
                    is_highlight_in_top_1000 = True
                    break
        
        filtered = filtered[:1000]

        # 하이라이트할 ID가 있고, 그 ID가 1000위 안에 없다면 리스트 맨 아래 추가
        if highlight_id and not is_highlight_in_top_1000:
            for r in full_sorted_list:
                if str(r['id']) == highlight_id:
                    # 순위권 밖 데이터 표시 설정
                    r['display_rank'] = "1000+"
                    r['is_outside_rank'] = True
                    filtered.append(r)
                    break

    return render_template(
        "rank.html",
        rows=filtered,
        current_range=current_range,
        current_level=current_level,
        current_size=current_size,
        levels=["all", "Kids", "Summer", "Space", "Santa", "Basic", "Easy", "Normal", "Hard", "Challenger"],
        sizes=["all", "소형", "중형", "대형"],
        LEVEL_KR=LEVEL_KR
    )


# ========================================================
# Web - Admin
# ========================================================
@web.route("/admin")
def admin_page():
    search_date = request.args.get("date")

    # 기본값 = 오늘
    if not search_date:
        search_date = datetime.now().strftime("%Y-%m-%d")

    rows = load_rows_by_date(search_date)
    
    return render_template(
        "admin.html",
        rows=rows,
        now=datetime.now(),
        search_date=search_date,
        sizes=["소형", "중형", "대형"],
        levels=["Kids", "Summer", "Space", "Santa", "Basic", "Easy", "Normal", "Hard", "Challenger"],
        pad_map=PAD_MAP,
        log_stats=log_stats,
        pad_status=pad_status,
        pad_data=pad_data
    )

@web.route("/admin/add", methods=["POST"])
def admin_add():
    dt = datetime(
        int(request.form["year"]),
        int(request.form["month"]),
        int(request.form["day"]),
        int(request.form["hour"]),
        int(request.form["minute"]),
    )

    save_to_db({
        "time": dt.strftime("%Y-%m-%d %H:%M"),
        "pad_id": int(request.form["pad_id"]),
        "map_id": None,
        "size": request.form["size"],
        "level": request.form["level"],
        "team": request.form["team"],
        "score": int(request.form["score"]),
        "rank": None,
    })

    print(request.form)
    return redirect(url_for("admin_page"))

@web.route("/admin/update/<int:id>", methods=["POST"])
def admin_update(id):
    time_val = request.form.get("time")
    size = request.form.get("size")
    level = request.form.get("level")
    team = request.form.get("team")
    score = request.form.get("score")

    # 🔒 방어 코드 (빈 값 방지)
    if not score:
        score = None
    else:
        score = int(score)

    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    cur.execute("""
        UPDATE game_records
        SET time=?, size=?, level=?, team=?, score=?
        WHERE id=?
    """, (time_val, size, level, team, score, id))

    updated = cur.rowcount > 0
    
    conn.commit()
    conn.close()

    if updated:
        print("🛠 관리자 수정 → 랭킹 리프레시 emit")
        socketio.emit("rank_refresh")

    print(request.form)
    return redirect(url_for("admin_page"))

@web.route("/admin/delete/<int:id>")
def admin_delete(id):
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    cur.execute("DELETE FROM game_records WHERE id=?", (id,))

    deleted = cur.rowcount > 0
    
    conn.commit()
    conn.close()

    if deleted:
        print("🛠 관리자 삭제 → 랭킹 리프레시 emit")
        socketio.emit("rank_refresh")
    
    return redirect(url_for("admin_page"))

@web.route("/team_search")
def team_search():
    keyword = request.args.get("keyword", "").replace(" ", "").lower()

    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    cur.execute("""
        SELECT id, time, size, level, team, score
        FROM game_records
        ORDER BY time DESC
    """)

    rows = []
    for r in cur.fetchall():
        team = r["team"] or ""
        team_norm = team.replace(" ", "").lower()

        if keyword in team_norm:
            rows.append({
                "id": r["id"],
                "rank": "-",
                "level": LEVEL_KR.get(r["level"], r["level"]),
                "raw_level": r["level"],
                "size": r["size"],
                "team": team,
                "score": r["score"],
                "time": r["time"]
            })

    conn.close()
    return jsonify(rows)


TEAM_LIST_SESSION_KEY = 'team_list_authenticated'


def team_list_api_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if session.get(TEAM_LIST_SESSION_KEY):
            return view(*args, **kwargs)
        return jsonify({'success': False, 'message': '직원 로그인이 필요합니다.'}), 401
    return wrapped


@web.route('/team_list', methods=['GET'])
def team_list_page():
    if not os.getenv('TEAM_LIST_PASSWORD'):
        return 'TEAM_LIST_PASSWORD 설정이 필요합니다.', 503
    if not session.get(TEAM_LIST_SESSION_KEY):
        return render_template('team_list_login.html')
    return render_template('team_list.html')


@web.route('/team_list/login', methods=['POST'])
def team_list_login():
    expected_password = os.getenv('TEAM_LIST_PASSWORD', '')
    supplied_password = request.form.get('password', '')
    if expected_password and hmac.compare_digest(supplied_password, expected_password):
        session[TEAM_LIST_SESSION_KEY] = True
        return redirect(url_for('team_list_page'))
    return render_template('team_list_login.html', error='비밀번호가 올바르지 않습니다.'), 401


@web.route('/team_list/logout', methods=['POST'])
def team_list_logout():
    session.pop(TEAM_LIST_SESSION_KEY, None)
    return redirect(url_for('team_list_page'))


@web.route('/api/teams')
@team_list_api_required
def get_teams():
    # 1. 프론트엔드에서 보낸 날짜 파라미터를 읽습니다. (없으면 오늘 날짜)
    target_date = request.args.get('date') or datetime.now().strftime('%Y-%m-%d')

    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    
    # 2. [핵심 수정] WHERE booking_date = ? 절을 추가하여 해당 날짜 데이터만 가져옵니다.
    cur.execute("""
        SELECT * FROM bookings 
        WHERE booking_date = ? 
        ORDER BY time_key ASC, room ASC, COALESCE(order_no, 0) ASC, id ASC
    """, (target_date,))
    
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()

    result = []
    for r in rows:
        room = (r.get('room') or '').upper()
        # 룸 이름에 따른 사이즈 자동 분류 (B는 중형, C는 소형)
        room_size = '중형' if room.startswith('B') else ('소형' if room.startswith('C') else '-')
        
        # payment_data 파싱 (전화번호가 payment_data 안에 있을 경우를 대비)
        p_data = _parse_payment_data_safe(r.get('payment_data'))
        
        result.append({
            'room': room or '-',
            'name': r.get('name') or '-',
            'team': r.get('team') or '개인',
            'level': r.get('level') or '-',
            'people': r.get('people') or '-',
            'room_size': room_size,
            'phone': r.get('phone') or '-',
            'is_paid': bool(r.get('paid')),
            'game_status': '✅완료' if bool(r.get('completed')) else '⏳진행중'
        })

    return jsonify(result)


# ========================================================
# Web - Dashboard
# ========================================================
@web.route("/dashboard")
def dashboard():
    # 1. 오늘 날짜를 생성해서 페이지로 보냅니다.
    #    이렇게 하면 HTML에서 "2026년 02월 11일 (수)" 처럼 표시할 수 있습니다.
    today_date = datetime.now().strftime('%Y-%m-%d')
    
    # 요일 표시를 위한 한국어 변환 (선택 사항)
    weekday_map = ['월', '화', '수', '목', '금', '토', '일']
    today_weekday = weekday_map[datetime.now().weekday()]
    
    # 2. 만약 DB가 있다면 여기서 데이터를 가져와서 template에 넘겨줍니다.
    # 현재는 빈 페이지를 띄우는 기본 설정입니다.
    template_path = os.path.join(web.root_path, "templates", "dashboard.html")
    with open(template_path, "r", encoding="utf-8") as template_file:
        template_source = template_file.read()

    return render_template_string(
        template_source,
        date_str=today_date,
        weekday=today_weekday,
    )

@web.route('/api/dashboard/update/<int:row_id>', methods=['POST'])
def update_data(row_id):
    # 1. 수정하려는 데이터의 기존 날짜 조회
    # 2. 오늘 날짜와 비교
    today = datetime.now().date()
    target_data = db.get_data_by_id(row_id)
    target_date = datetime.strptime(target_data['date'], '%Y-%m-%d').date()

    if target_date < today:
        return jsonify({"message": "과거 데이터는 수정할 수 없습니다."}), 403

    # 수정 로직 진행...
    return jsonify({"message": "수정 완료"})


@web.route('/settlement')
def settlement_page():
    return render_template('settlement.html')


@web.route('/api/settlement/overview', methods=['GET'])
def get_settlement_overview():
    target_date = _get_request_date(default_today=True)
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    cur.execute(
        "SELECT id, booking_date, time_key, room, name, phone, team, people, paid, completed, payment_data FROM bookings WHERE booking_date=? ORDER BY time_key, room, COALESCE(order_no, 0), id",
        (target_date,)
    )
    booking_rows = [dict(r) for r in cur.fetchall()]

    cur.execute(
        "SELECT id, target_date, time, item, quantity, etc_text, card_amount, cash_amount, transfer_amount, total_amount FROM supply_history WHERE target_date=? ORDER BY id ASC",
        (target_date,)
    )
    supply_rows = [dict(r) for r in cur.fetchall()]

    cur.execute(
        "SELECT cash_expense, no_show_count FROM settlement_daily_meta WHERE target_date=?",
        (target_date,)
    )
    meta_row = cur.fetchone()
    cash_expense = int(meta_row['cash_expense'] or 0) if meta_row else 0
    manual_no_show_count = int(meta_row['no_show_count'] or 0) if meta_row else 0
    manual_no_show_amount = manual_no_show_count * 5000
    conn.close()

    team_rows = []
    team_card = 0
    team_cash = 0
    team_transfer = 0

    for row in booking_rows:
        payment_data = _parse_payment_data_safe(row.get('payment_data'))
        card_amount = int(payment_data.get('cardInput') or 0)
        cash_amount = int(payment_data.get('cashInput') or 0)
        transfer_amount = int(payment_data.get('transferInput') or 0)
        total_amount = card_amount + cash_amount + transfer_amount

        team_rows.append({
            'id': row['id'],
            'booking_date': row.get('booking_date'),
            'time_key': row.get('time_key'),
            'room': row.get('room'),
            'name': row.get('name') or '',
            'team': row.get('team') or '',
            'people': row.get('people') or '',
            'paid': int(row.get('paid') or 0),
            'completed': int(row.get('completed') or 0),
            'card_amount': card_amount,
            'cash_amount': cash_amount,
            'transfer_amount': transfer_amount,
            'total_amount': total_amount,
            'payment_data': payment_data,
        })

        if int(row.get('paid') or 0):
            team_card += card_amount
            team_cash += cash_amount
            team_transfer += transfer_amount

    supply_card = sum(int(r.get('card_amount') or 0) for r in supply_rows)
    supply_cash = sum(int(r.get('cash_amount') or 0) for r in supply_rows)
    supply_transfer = sum(int(r.get('transfer_amount') or 0) for r in supply_rows)

    no_show_total = 0
    supply_deposit_total = 0
    supply_card_excluding_deposit = 0
    for r in supply_rows:
        item_text = str(r.get('item') or '').replace(' ', '')
        amount = int(r.get('card_amount') or 0) + int(r.get('cash_amount') or 0) + int(r.get('transfer_amount') or 0)
        is_deposit_entry = '(예)' in item_text
        if is_deposit_entry:
            supply_deposit_total += amount
        else:
            supply_card_excluding_deposit += int(r.get('card_amount') or 0)
        if ('당일취소' in item_text) or ('노쇼' in item_text) or ('취소&노쇼' in item_text):
            no_show_total += amount

    team_deposit_total = 0
    for row in team_rows:
        pd = row.get('payment_data') or {}
        if pd.get('depositPaid'):
            team_deposit_total += int(pd.get('depositAmount') or 5000)

    deposit_total = team_deposit_total + supply_deposit_total + manual_no_show_amount
    combined_card = team_card + supply_card_excluding_deposit
    combined_cash_before_expense = team_cash + supply_cash
    combined_cash = combined_cash_before_expense - cash_expense
    combined_transfer = team_transfer + supply_transfer
    display_no_show_total = no_show_total + manual_no_show_amount
    combined_total = combined_card + combined_cash + combined_transfer + deposit_total + no_show_total

    return jsonify({
        'target_date': target_date,
        'team_rows': team_rows,
        'supply_rows': supply_rows,
        'totals': {
            'team': {
                'card': team_card,
                'cash': team_cash,
                'transfer': team_transfer,
                'total': team_card + team_cash + team_transfer,
            },
            'supply': {
                'card': supply_card,
                'cash': supply_cash,
                'transfer': supply_transfer,
                'total': supply_card + supply_cash + supply_transfer,
            },
            'combined': {
                'card': combined_card,
                'cash': combined_cash,
                'cash_before_expense': combined_cash_before_expense,
                'cash_expense': cash_expense,
                'no_show_count': manual_no_show_count,
                'transfer': combined_transfer,
                'deposit': deposit_total,
                'no_show': display_no_show_total,
                'total': combined_total,
            }
        }
    })


@web.route('/api/settlement/cash_expense', methods=['GET'])
def get_settlement_cash_expense():
    target_date = _get_request_date(default_today=True)
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute(
        "SELECT cash_expense FROM settlement_daily_meta WHERE target_date=?",
        (target_date,)
    )
    row = cur.fetchone()
    conn.close()
    amount = int(row['cash_expense'] or 0) if row else 0
    return jsonify({'target_date': target_date, 'cash_expense': amount})


@web.route('/api/settlement/cash_expense', methods=['PUT'])
def save_settlement_cash_expense():
    target_date = _get_request_date(default_today=True)
    data = request.json or {}
    cash_expense = max(int(data.get('cashExpense') or 0), 0)

    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    cur.execute(
        '''INSERT INTO settlement_daily_meta (target_date, cash_expense, updated_at)
           VALUES (?, ?, datetime('now', 'localtime'))
           ON CONFLICT(target_date) DO UPDATE SET
             cash_expense=excluded.cash_expense,
             updated_at=datetime('now', 'localtime')''',
        (target_date, cash_expense)
    )
    conn.commit()
    conn.close()
    return jsonify({'status': 'success', 'target_date': target_date, 'cash_expense': cash_expense})


@web.route('/api/settlement/no_show', methods=['GET'])
def get_settlement_no_show_count():
    target_date = _get_request_date(default_today=True)
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute(
        "SELECT no_show_count FROM settlement_daily_meta WHERE target_date=?",
        (target_date,)
    )
    row = cur.fetchone()
    conn.close()
    no_show_count = int(row['no_show_count'] or 0) if row else 0
    return jsonify({'target_date': target_date, 'no_show_count': no_show_count})


@web.route('/api/settlement/no_show', methods=['PUT'])
def save_settlement_no_show_count():
    target_date = _get_request_date(default_today=True)
    data = request.json or {}
    no_show_count = max(int(data.get('noShowCount') or 0), 0)

    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    cur.execute(
        '''INSERT INTO settlement_daily_meta (target_date, no_show_count, updated_at)
           VALUES (?, ?, datetime('now', 'localtime'))
           ON CONFLICT(target_date) DO UPDATE SET
             no_show_count=excluded.no_show_count,
             updated_at=datetime('now', 'localtime')''',
        (target_date, no_show_count)
    )
    conn.commit()
    conn.close()
    return jsonify({'status': 'success', 'target_date': target_date, 'no_show_count': no_show_count})


@web.route('/api/settlement/team/<int:bid>', methods=['PUT'])
def update_settlement_team_payment(bid):
    data = request.json or {}
    time_key = str(data.get('time_key', '')).strip() or '00-00'
    room = str(data.get('room', '')).strip() or 'C1'
    name = str(data.get('name', '')).strip()
    team = str(data.get('team', '')).strip()
    people = max(int(data.get('people') or 0), 0)
    total_people = max(int(data.get('total_people') or people), 0)
    adult_count = max(int(data.get('adult_count') or 0), 0)
    child_count = max(int(data.get('child_count') or 0), 0)
    deposit_amount = max(int(data.get('deposit_amount') or 0), 0)
    pass_adult_count = max(int(data.get('pass_adult_count') or 0), 0)
    pass_child_count = max(int(data.get('pass_child_count') or 0), 0)
    coupon_count = max(int(data.get('coupon_count') or 0), 0)
    card_amount = max(int(data.get('card_amount') or 0), 0)
    cash_amount = max(int(data.get('cash_amount') or 0), 0)
    transfer_amount = max(int(data.get('transfer_amount') or 0), 0)
    paid = 1 if bool(data.get('paid')) else 0

    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute('SELECT paid, payment_data FROM bookings WHERE id=?', (bid,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return jsonify({'status': 'error', 'message': 'Booking not found'}), 404

    payment_data = _parse_payment_data_safe(row['payment_data'])
    payment_data['cardInput'] = card_amount
    payment_data['cashInput'] = cash_amount
    payment_data['transferInput'] = transfer_amount
    payment_data['isMatching'] = bool(paid)
    payment_data['totalPeople'] = total_people
    payment_data['adultCount'] = adult_count
    payment_data['childCount'] = child_count
    payment_data['adultPass'] = pass_adult_count
    payment_data['childPass'] = pass_child_count
    payment_data['coupon'] = coupon_count
    payment_data['depositPaid'] = bool(deposit_amount > 0)
    payment_data['depositAmount'] = deposit_amount

    if 'finalPaymentAmount' not in payment_data:
        payment_data['finalPaymentAmount'] = card_amount + cash_amount + transfer_amount

    cur.execute(
        'UPDATE bookings SET time_key=?, room=?, name=?, team=?, people=?, paid=?, payment_data=? WHERE id=?',
        (time_key, room, name, team, str(people or total_people or ''), paid, json.dumps(payment_data, ensure_ascii=False), bid)
    )
    conn.commit()
    conn.close()

    return jsonify({'status': 'success', 'id': bid})


@web.route('/api/settlement/team', methods=['POST'])
def create_settlement_team():
    data = request.json or {}
    booking_date = str(data.get('booking_date', '')).strip()
    if not booking_date:
        return jsonify({'status': 'error', 'message': 'booking_date required'}), 400

    time_key = str(data.get('time_key', '')).strip() or '00-00'
    room = str(data.get('room', '')).strip() or '-'
    name = str(data.get('name', '')).strip()
    team = str(data.get('team', '')).strip()
    phone = str(data.get('phone', '')).strip()
    people = max(int(data.get('people') or 0), 0)
    total_people = max(int(data.get('total_people') or people), 0)
    adult_count = max(int(data.get('adult_count') or 0), 0)
    child_count = max(int(data.get('child_count') or 0), 0)
    deposit_amount = max(int(data.get('deposit_amount') or 0), 0)
    pass_adult_count = max(int(data.get('pass_adult_count') or 0), 0)
    pass_child_count = max(int(data.get('pass_child_count') or 0), 0)
    coupon_count = max(int(data.get('coupon_count') or 0), 0)
    card_amount = max(int(data.get('card_amount') or 0), 0)
    cash_amount = max(int(data.get('cash_amount') or 0), 0)
    transfer_amount = max(int(data.get('transfer_amount') or 0), 0)
    paid = 1 if bool(data.get('paid')) else 0

    payment_data = {
        'cardInput': max(int(data.get('card_amount') or 0), 0),
        'cashInput': max(int(data.get('cash_amount') or 0), 0),
        'transferInput': max(int(data.get('transfer_amount') or 0), 0),
        'isMatching': bool(paid),
        'totalPeople': people,
        'depositPaid': max(int(data.get('deposit_amount') or 0), 0) > 0,
        'depositAmount': max(int(data.get('deposit_amount') or 0), 0),
        'roomFlags': {"F": False, "S": False, "M": False, "L": False}, # 호환성 유지
        'finalPaymentAmount': max(int(data.get('card_amount') or 0), 0) + 
                                max(int(data.get('cash_amount') or 0), 0) + 
                                max(int(data.get('transfer_amount') or 0), 0)
    }

    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    cur.execute(
        'INSERT INTO bookings (booking_date, time_key, room, name, phone, team, people, paid, payment_data, created_at) VALUES (?,?,?,?,?,?,?,?,?,datetime("now","localtime"))',
        (booking_date, time_key, room, name, phone, team, str(people or total_people or ''), paid, json.dumps(payment_data, ensure_ascii=False))
    )
    new_id = cur.lastrowid
    conn.commit()
    conn.close()

    return jsonify({'status': 'ok', 'id': new_id})

@web.route('/api/settlement/team/<int:bid>', methods=['DELETE'])
def delete_settlement_team(bid):
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    try:
        # 데이터가 있는지 확인
        cur.execute('SELECT id FROM bookings WHERE id=?', (bid,))
        if not cur.fetchone():
            conn.close()
            return jsonify({'status': 'error', 'message': '이미 삭제되었거나 존재하지 않는 데이터입니다.'}), 404

        # 삭제 수행
        cur.execute('DELETE FROM bookings WHERE id=?', (bid,))
        conn.commit()
        return jsonify({'status': 'success', 'message': '정상적으로 삭제되었습니다.'})
    except Exception as e:
        print(f"삭제 중 오류 발생: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        conn.close()


@web.route('/api/day_type_override', methods=['GET'])
def get_day_type_override():
    target_date = _get_request_date(default_today=True)
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute('SELECT day_type FROM day_type_overrides WHERE target_date=?', (target_date,))
    row = cur.fetchone()
    conn.close()
    return jsonify({
        'target_date': target_date,
        'day_type': row['day_type'] if row else None
    })


@web.route('/api/day_type_override', methods=['PUT'])
def put_day_type_override():
    data = request.json or {}
    target_date = _normalize_date_str(data.get('target_date'))
    day_type = str(data.get('day_type', '')).strip()
    if not target_date:
        return jsonify({'status': 'error', 'message': 'Invalid target_date'}), 400
    if day_type not in {'weekday', 'weekend'}:
        return jsonify({'status': 'error', 'message': 'Invalid day_type'}), 400

    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    cur.execute(
        '''INSERT INTO day_type_overrides (target_date, day_type, updated_at)
           VALUES (?, ?, CURRENT_TIMESTAMP)
           ON CONFLICT(target_date) DO UPDATE SET
             day_type=excluded.day_type,
             updated_at=CURRENT_TIMESTAMP''',
        (target_date, day_type)
    )
    conn.commit()
    conn.close()
    return jsonify({'status': 'success', 'target_date': target_date, 'day_type': day_type})


@web.route('/api/day_type_override', methods=['DELETE'])
def delete_day_type_override():
    target_date = _get_request_date(default_today=False)
    if not target_date:
        data = request.json or {}
        target_date = _normalize_date_str(data.get('target_date'))
    if not target_date:
        return jsonify({'status': 'error', 'message': 'Invalid target_date'}), 400

    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    cur.execute('DELETE FROM day_type_overrides WHERE target_date=?', (target_date,))
    conn.commit()
    conn.close()
    return jsonify({'status': 'success', 'target_date': target_date})



# ========================================================
# Web - Booking
# ========================================================
@web.route('/api/booking/add', methods=['POST'])
def add_booking():
    data = request.json or {}
    
    # 1. 필수 값 검증 (기존 로직 유지)
    if not all(k in data for k in ['time_key', 'room']):
        return jsonify({"status": "error", "message": "Missing time_key or room"}), 400

    booking_date = _normalize_date_str(data.get('booking_date')) or _today_date_str()
    team = str(data.get('team', '')).rstrip()[:10]
    order_no = data.get('order_no')

    # 🎯 [교정 1]: 타임아웃 30초 장착으로 줄 서서 대기하게 만듭니다.
    conn = sqlite3.connect(DB_FILE, timeout=30)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()


    try:
        # 🎯 [교정 2]: 순번 조회(SELECT)부터 등록(INSERT)까지 통째로 하나의 보호막(with conn) 안에 가둡니다.
        # 이렇게 해야 중간에 메일 감시망이 새치기를 못 합니다.
        with conn:
            # 순번 번호가 없을 때 자동 계산 기능 (기존 로직 유지)
            if order_no is None:
                cur.execute(
                    'SELECT COALESCE(MAX(order_no), 0) + 1 FROM bookings WHERE time_key=? AND room=?',
                    (data['time_key'], data['room'])
                )
                order_no = cur.fetchone()[0]

            # 데이터 전격 등록
            cur.execute(
                '''INSERT INTO bookings
                   (booking_date, time_key, room, name, phone, team, level, people, order_no, paid, completed, payment_data, created_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)''',
                (
                    booking_date, data['time_key'], data['room'],
                    data.get('name', ''), data.get('phone', ''), team, data.get('level', ''), data.get('people', ''), int(order_no),
                    data.get('paid', 0), data.get('completed', 0), data.get('payment_data', None)
                )
            )
            new_id = cur.lastrowid
            naver_booking_id = str(data.get('naver_booking_id', '')).strip()
            if naver_booking_id:
                cur.execute(
                    '''INSERT INTO naver_booking_card_links (booking_id, booking_row_id, handling_mode, card_state)
                       VALUES (?, ?, 'standard', 'active')
                       ON CONFLICT(booking_id) DO UPDATE SET
                           booking_row_id=excluded.booking_row_id,
                           card_state='active',
                           updated_at=CURRENT_TIMESTAMP''',
                    (naver_booking_id[:100], new_id),
                )

        # 🚀 [교정 3]: with 블록을 빠져나오면 자동으로 안전하게 Commit 완료됩니다!
        return jsonify({"status": "success", "id": new_id})

    except Exception as e:
        # 사장님이 구축하신 일자별 파일 로그 시스템에 기록되도록 연동
        # .logger.error(f"❌ [예약 신규 등록 실패] 에러 원인: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500

    finally:
        # 🔒 에러가 나든 성공하든 디비 통로는 칼같이 반납
        cur.close()
        conn.close()


@web.route('/api/booking/<int:bid>', methods=['PUT'])
def update_booking(bid):
    data = request.json or {}
    conn = sqlite3.connect(DB_FILE, timeout=30)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute('SELECT booking_date FROM bookings WHERE id=?', (bid,))
    existing = cur.fetchone()
    if not existing:
        conn.close()
        return jsonify({"status": "error", "message": "Booking not found"}), 404

    booking_date = _normalize_date_str(data.get('booking_date')) or existing['booking_date'] or _today_date_str()
    team = str(data.get('team', '')).rstrip()[:10]
    try:
        with conn:
            cur.execute(
            '''UPDATE bookings
            SET booking_date=?, time_key=?, room=?, name=?, phone=?, team=?, level=?, people=?, order_no=?, paid=?, completed=?, payment_data=?
            WHERE id=?''',
            (
                booking_date, data.get('time_key'), data.get('room'),
                data.get('name', ''), data.get('phone', ''), team, data.get('level', ''), data.get('people', ''), int(data.get('order_no', 0) or 0),
                data.get('paid', 0), data.get('completed', 0), data.get('payment_data', None), bid
            )
        )
        return jsonify({"success": True, "message": "수정 완료"})
        
    except Exception as e:
        print(f"❌ [예약 수정 실패] 에러: {e}")
        return jsonify({"success": False, "error": str(e)})
        
    finally:
        cur.close()
        conn.close()
    
    


@web.route('/api/booking/<int:bid>', methods=['DELETE'])
def delete_booking(bid):
    conn = sqlite3.connect(DB_FILE, timeout=30)
    cur = conn.cursor()
    try:
        # 🎯 [교정 2]: with conn 보호막을 씌워 데이터 삭제 즉시 자동으로 Commit 하고 잠금을 풀게 합니다.
        with conn:
            cur.execute('DELETE FROM bookings WHERE id=?', (bid,))
            
        # 🚀 디비에서 완전히 증발한 것을 확인한 뒤 프론트엔드에 성공 리턴
        return jsonify({"status": "success"})
        
    except Exception as e:
        # 사장님의 일자별 파일 로그 시스템에 에러 기록
        web.logger.error(f"❌ [예약 삭제 실패] 에러 원인: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500
        
    finally:
        # 🔒 어떤 예외 상황이 오든 자원은 무조건 정갈하게 반납!
        cur.close()
        conn.close()


@web.route('/api/booking/list', methods=['GET'])
def get_booking_list():
    target_date = _get_request_date(default_today=True)
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute(
        '''SELECT * FROM bookings
             WHERE booking_date=?
               AND NOT EXISTS (
                   SELECT 1 FROM naver_booking_card_links AS link
                    WHERE link.booking_row_id=bookings.id
                      AND link.card_state='cancelled_hidden'
               )
             ORDER BY time_key, room, COALESCE(order_no, 0), id''',
        (target_date,),
    )
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return jsonify(rows)


def _time_key_sort_value(time_key):
    try:
        hour_str, minute_str = str(time_key).split('-', 1)
        return int(hour_str), int(minute_str)
    except Exception:
        return (99, 99)


def _format_time_key(time_key):
    hour, minute = _time_key_sort_value(time_key)
    if hour == 99:
        return str(time_key)
    return f"{hour:02d}:{minute:02d}"


def _build_booking_cell_text(row):
    team = (row.get('team') or '').strip()
    name = (row.get('name') or '').strip()
    phone = (row.get('phone') or '').strip()
    level = (row.get('level') or '').strip()
    people = str(row.get('people') or '').strip()
    status = '결완' if int(row.get('paid') or 0) else '미완'

    first_line = team or name or '-'
    extras = [value for value in [name if team and name and name != team else '', level, people] if value]
    second_line = ' / '.join(extras)
    return f"{first_line}\n{second_line}\n{status}" if second_line else f"{first_line}\n{status}"


@web.route('/api/booking/export-excel', methods=['GET'])
def export_booking_excel():
    target_date = _get_request_date(default_today=True)
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT * FROM bookings WHERE booking_date=? ORDER BY time_key, room, COALESCE(order_no, 0), id", (target_date,))
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = '타임테이블'

    header_fill = PatternFill('solid', fgColor='DCE6F1')
    room_fill = PatternFill('solid', fgColor='EAF4EA')
    thin_side = Side(style='thin', color='B7C0CC')
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    top_align = Alignment(vertical='top', wrap_text=True)
    bold_font = Font(bold=True)

    rooms = ['C1', 'C2', 'B1', 'B2']
    sheet.append(['TIME', *rooms])
    for cell in sheet[1]:
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = cell_border

    grouped = {}
    for row in rows:
        grouped.setdefault((row['time_key'], row['room']), []).append(row)

    time_keys = sorted({row['time_key'] for row in rows}, key=_time_key_sort_value)
    if not time_keys:
        time_keys = [f"{hour}-{minute}" for hour in range(10, 23) for minute in (0, 20, 40)]

    for time_key in time_keys:
        row_values = [_format_time_key(time_key)]
        for room in rooms:
            cell_rows = grouped.get((time_key, room), [])
            row_values.append('\n\n'.join(_build_booking_cell_text(item) for item in cell_rows))
        sheet.append(row_values)

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        sheet.row_dimensions[row_idx].height = 54
        for col_idx, cell in enumerate(row, start=1):
            cell.border = cell_border
            if col_idx == 1:
                cell.font = bold_font
                cell.fill = header_fill
                cell.alignment = center_align
            else:
                cell.alignment = top_align

    for col_letter, width in {'A': 11, 'B': 24, 'C': 24, 'D': 24, 'E': 24}.items():
        sheet.column_dimensions[col_letter].width = width

    detail_sheet = workbook.create_sheet('예약목록')
    detail_headers = ['시간', '룸', '팀명', '성함', '난이도', '인원', '결제상태', '완료', '결제정보']
    detail_sheet.append(detail_headers)
    for cell in detail_sheet[1]:
        cell.font = bold_font
        cell.fill = room_fill
        cell.alignment = center_align
        cell.border = cell_border

    for row in sorted(rows, key=lambda item: (_time_key_sort_value(item['time_key']), item['room'], int(item.get('order_no') or 0), int(item.get('id') or 0))):
        payment_summary = ''
        raw_payment_data = row.get('payment_data')
        if raw_payment_data:
            try:
                payment_data = json.loads(raw_payment_data)
                parts = []
                if payment_data.get('cardInput'):
                    parts.append(f"카드 {int(payment_data['cardInput']):,}원")
                if payment_data.get('cashInput'):
                    parts.append(f"현금 {int(payment_data['cashInput']):,}원")
                if payment_data.get('transferInput'):
                    parts.append(f"계좌 {int(payment_data['transferInput']):,}원")
                if payment_data.get('depositAmount'):
                    parts.append(f"예약금 {int(payment_data['depositAmount']):,}원")
                if payment_data.get('partyRoom'):
                    parts.append('파티룸')
                payment_summary = ' / '.join(parts)
            except Exception:
                payment_summary = str(raw_payment_data)

        detail_sheet.append([
            _format_time_key(row.get('time_key')),
            row.get('room', ''),
            row.get('team', ''),
            row.get('name', ''),
            row.get('level', ''),
            row.get('people', ''),
            '결제완료' if int(row.get('paid') or 0) else '결제미완료',
            '완료' if int(row.get('completed') or 0) else '',
            payment_summary,
        ])

    for row in detail_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = cell_border
            cell.alignment = top_align

    for col_letter, width in {'A': 10, 'B': 8, 'C': 16, 'D': 12, 'E': 12, 'F': 10, 'G': 12, 'H': 8, 'I': 36}.items():
        detail_sheet.column_dimensions[col_letter].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"jumpingbattle_timetable_{target_date.replace('-', '')}_{datetime.now().strftime('%H%M')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@web.route('/api/queue/list', methods=['GET'])
def get_queue_list():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    ensure_queue_items_order_column(conn)
    cur = conn.cursor()
    cur.execute("SELECT * FROM queue_items ORDER BY room, COALESCE(order_no, id), id")
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return jsonify(rows)


@web.route('/api/queue/add', methods=['POST'])
def add_queue_item():
    data = request.json or {}
    room = str(data.get('room', '')).strip().upper()
    if room not in {'C1', 'C2', 'B1', 'B2'}:
        return jsonify({"status": "error", "message": "Invalid room"}), 400

    conn = sqlite3.connect(DB_FILE)
    ensure_queue_items_order_column(conn)
    cur = conn.cursor()
    cur.execute('SELECT COALESCE(MAX(order_no), 0) FROM queue_items WHERE room=?', (room,))
    next_order_no = (cur.fetchone()[0] or 0) + 1
    cur.execute(
        'INSERT INTO queue_items (room, order_no, name, team, level, people, bid, party_room) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
        (
            room,
            next_order_no,
            str(data.get('name', '')).strip(),
            str(data.get('team', '')).rstrip()[:10],
            str(data.get('level', '')).strip(),
            str(data.get('people', '')).strip(),
            int(data.get('bid') or 0),
            1 if data.get('partyRoom') else 0,
        )
    )
    new_id = cur.lastrowid
    conn.commit()
    conn.close()
    # 대기리스트 변경 시 대시보드에 알림
    socketio.emit('room_or_queue_changed')
    return jsonify({"status": "success", "id": new_id})


# ========================================================
# Web - Reservation from Naver Booking
# ========================================================

@web.route('/api/naver-reservations', methods=['GET'])
@team_list_api_required
def get_naver_reservation_list():
    target_date = (request.args.get('date') or datetime.now().strftime('%Y-%m-%d')).strip()
    try:
        datetime.strptime(target_date, '%Y-%m-%d')
    except ValueError:
        return jsonify({'success': False, 'message': '날짜 형식이 올바르지 않습니다.'}), 400

    status_labels = {
        'CONFIRMED': '예약',
        'COMPLETED': '이용 완료',
        'CANCELED': '취소',
    }
    with get_db_connection() as conn:
        rows = conn.execute(
            '''
            SELECT reservation.booking_id, reservation.booking_status, reservation.use_time_key, reservation.room_name, reservation.product_name,
                   customer_name, team_name, difficulty, phone, people_count,
                   first_seen_at, reservation.updated_at, cancelled_at,
                   COALESCE(link.handling_mode, 'standard') AS handling_mode
              FROM naver_reservations AS reservation
              LEFT JOIN naver_booking_card_links AS link ON link.booking_id=reservation.booking_id
             WHERE reservation.use_date=?
             ORDER BY reservation.use_time_key ASC, reservation.room_name ASC, reservation.first_seen_at ASC
            ''',
            (target_date,),
        ).fetchall()

    return jsonify([
        {
            'booking_id': row[0],
            'status': row[1],
            'status_label': '현장 결제 전환' if row[13] == 'onsite_payment' else status_labels.get(row[1], row[1]),
            'handling_mode': row[13],
            'time': str(row[2] or '').replace('-', ':'),
            'room': row[3] or '-',
            'product': row[4] or '-',
            'name': row[5] or '-',
            'team': row[6] or '-',
            'difficulty': row[7] or '-',
            'phone': row[8] or '-',
            'people': '' if not row[9] else row[9],
            'received_at': row[10] or '',
            'updated_at': row[11] or '',
            'cancelled_at': row[12] or '',
        }
        for row in rows
    ])


@web.route('/api/naver-reservations/<booking_id>/onsite-payment', methods=['POST'])
@team_list_api_required
def convert_naver_reservation_to_onsite_payment(booking_id):
    booking_id = str(booking_id or '').strip()
    if not booking_id or len(booking_id) > 100:
        return jsonify({'success': False, 'message': '예약번호가 올바르지 않습니다.'}), 400

    with get_db_connection() as conn:
        reservation = conn.execute(
            "SELECT booking_status FROM naver_reservations WHERE booking_id=?",
            (booking_id,),
        ).fetchone()
        link = conn.execute(
            "SELECT booking_row_id FROM naver_booking_card_links WHERE booking_id=?",
            (booking_id,),
        ).fetchone()
        if not reservation:
            return jsonify({'success': False, 'message': '예약 원본을 찾을 수 없습니다.'}), 404
        if reservation[0] == 'CANCELED':
            return jsonify({'success': False, 'message': '이미 취소된 예약입니다. 카드 복구 처리가 필요합니다.'}), 409
        if not link or not link[0]:
            return jsonify({'success': False, 'message': '시간표에 게임 카드가 등록된 뒤 전환할 수 있습니다.'}), 409

        booking = conn.execute(
            "SELECT payment_data FROM bookings WHERE id=?",
            (link[0],),
        ).fetchone()
        if not booking:
            return jsonify({'success': False, 'message': '연결된 게임 카드를 찾을 수 없습니다.'}), 409
        try:
            payment_data = json.loads(booking[0] or '{}')
        except (TypeError, json.JSONDecodeError):
            payment_data = {}
        payment_data.update({
            'isBooker': True,
            'depositPaid': False,
            'depositAmount': 0,
            'onsitePayment': True,
            'naverBookingId': booking_id,
        })
        conn.execute(
            "UPDATE bookings SET payment_data=? WHERE id=?",
            (json.dumps(payment_data, ensure_ascii=False), link[0]),
        )
        conn.execute(
            '''UPDATE naver_booking_card_links
                  SET handling_mode='onsite_payment', card_state='active', updated_at=CURRENT_TIMESTAMP
                WHERE booking_id=?''',
            (booking_id,),
        )
    return jsonify({'success': True, 'message': '현장 결제 전환으로 처리했습니다.'})

@web.route('/api/naver-bookings/today-init', methods=['GET'])
def get_today_init_bookings():
    today_str = datetime.now().strftime('%Y-%m-%d') # 오늘 날짜 (2026-05-18)
    
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    # 오늘 이용 조건 + 상태가 INIT인 건만 조회
    query = """
        SELECT
            cache.booking_id,
            cache.masked_name,
            cache.use_time_key,
            cache.room_name,
            COALESCE(reservation.team_name, ''),
            COALESCE(reservation.difficulty, ''),
            COALESCE(reservation.phone, ''),
            COALESCE(reservation.people_count, '')
        FROM naver_mail_cache AS cache
        LEFT JOIN naver_reservations AS reservation ON reservation.booking_id = cache.booking_id
        WHERE cache.use_date = ? AND cache.status = 'INIT'
        ORDER BY cache.use_time_key ASC
    """
    cursor.execute(query, (today_str,))
    rows = cursor.fetchall()
    conn.close()
    
    # 프론트엔드(JS)가 쓰기 좋게 JSON 배열로 가공
    booking_list = []
    for row in rows:
        booking_list.append({
            'booking_id': row[0],
            'name': row[1],
            'time': row[2].replace('-', ':'),
            'room': row[3],
            'team': row[4],
            'difficulty': row[5],
            'phone': row[6],
            'people': row[7],
        })
        
    return jsonify(booking_list)


@web.route('/api/naver-bookings/confirm', methods=['POST'])
def confirm_naver_booking():
    data = request.json
    booking_id = data.get('booking_id')
    
    if not booking_id:
        return jsonify({'success': False, 'message': '예약번호 누락'}), 400
        
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    
    try:
        # ⭐️ [핵심] 기존에 존재하던 INSERT INTO bookings 관련 쿼리를 통째로 삭제했습니다.
        # 이제 백엔드는 대기 박스에서 카드가 사라지도록 상태값만 안전하게 업데이트합니다.
        cur.execute('''
            UPDATE naver_mail_cache 
            SET status = 'CONFIRMED' 
            WHERE booking_id = ?
        ''', (booking_id,))
        
        conn.commit()
        print(f"✅ [네이버 캐시 마감 완료] 예약번호 {booking_id} 상태 변경완료 (CONFIRMED)")
        return jsonify({'success': True, 'message': '상태 변경 완료'})

    except sqlite3.Error as e:
        conn.rollback()
        print(f"❌ [DB 에러] 캐시 상태 변경 실패: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        cur.close()
        conn.close()


@web.route('/api/naver-bookings/webhook', methods=['POST'])
def naver_booking_webhook():
    try:
        # 구글 웹훅에서 보낸 JSON 데이터 확보
        print("🔔 [구글 웹훅 온 신호] 무언가 신호가 들어왔습니다!!")
        data = request.json
        print(f"📦 [들어온 데이터 내용]: {data}")

        if not data or 'content' not in data:
            return jsonify({'success': False, 'message': '올바르지 않은 웹훅 데이터 형식입니다.'}), 400
        
        raw_email_content = data['content']
        
        # 💡 아까 완성한 메일 파싱 및 캐시 DB 저장 함수 호출!
        parse_result = parse_and_save_naver_email(raw_email_content)
        
        if parse_result:
            socketio.emit('walkin_added')
            print(f"✅ [실시간 푸시 발송 완료] 번호: {parse_result.get('res_id')}")
            return jsonify({'success': True, 'message': '네이버 예약 캐시 저장 및 실시간 푸시 완료', 'data': parse_result}), 200
        else:
            socketio.emit('walkin_added')
            return jsonify({'success': False, 'message': '중복 건이거나 취소/파싱 실패입니다.'}), 200

    except Exception as e:
        print(f"❌ [웹훅 라우트 에러] {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

def check_cancellations_on_startup():
    """
    💡 서버 구동 시 Gmail에 접속하여 안 읽은 네이버 예약 취소 메일을 싹 긁어와
    대시보드 DB(캐시 및 정식 예약)에서 해당 건들을 일괄 삭제하는 함수
    """
    print("🔍 [서버 구동] 네이버 예약 취소 메일 일괄 스캔을 시작합니다...")
    
    # Keep account credentials in the ignored local .env file.
    IMAP_SERVER = os.getenv("GMAIL_IMAP_SERVER", "imap.gmail.com")
    IMAP_USER = os.getenv("GMAIL_USER")
    IMAP_PASSWORD = os.getenv("GMAIL_APP_PASSWORD")
    if not IMAP_USER or not IMAP_PASSWORD:
        print("[예약 취소 동기화 건너뜀] Gmail 환경설정이 없습니다.")
        return

    try:
        # 1. Gmail IMAP 서버 연결
        mail = imaplib.IMAP4_SSL(IMAP_SERVER)
        mail.login(IMAP_USER, IMAP_PASSWORD)
        mail.select("inbox")

        # 2. 네이버 예약 취소 메일 중 '읽지 않은(UNSEEN)' 메일만 검색
        # 영문 기준 검색을 위해 포맷팅 규칙 적용
        search_query = '(FROM "booking-noreply@naver.com" UNSEEN)'
        status, messages = mail.search(None, search_query)
        
        if status != "OK" or not messages[0]:
            print("✨ [스캔 완료] 처리할 새로운 취소 메일이 없습니다.")
            mail.logout()
            return

        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()
        cancel_count = 0

        # 3. 검색된 메일들을 하나씩 파싱
        for num in messages[0].split():
            status, data = mail.fetch(num, '(RFC822)')
            if status != "OK":
                continue
                
            raw_email = data[0][1]
            msg = email.message_from_bytes(raw_email)
            
            # 메일 제목 추출 및 디코딩
            subject, encoding = decode_header(msg["Subject"])[0]
            if isinstance(subject, bytes):
                subject = subject.decode(encoding or "utf-8", errors="ignore")
            
            # 💡 제목에 '취소'라는 단어가 들어간 메일만 정밀 타격
            if "취소" in subject:
                # 메일 본문 추출
                body = ""
                if msg.is_multipart():
                    for part in msg.walk():
                        if part.get_content_type() == "text/plain":
                            body = part.get_payload(decode=True).decode("utf-8", errors="ignore")
                            break
                else:
                    body = msg.get_payload(decode=True).decode("utf-8", errors="ignore")

                # 본문에서 예약번호 추출
                booking_id_match = re.search(r'예약번호\s*([A-Z0-9-]+|[0-9]+)', body)
                booking_id = booking_id_match.group(1).strip() if booking_id_match else None

                if booking_id:
                    # A. 워크인 대기 캐시 테이블에서 삭제
                    cur.execute("DELETE FROM naver_mail_cache WHERE booking_id = ?", (booking_id,))
                    
                    # B. 정식 타임테이블(bookings)의 team 또는 name 컬럼에 해당 예약번호(이름)가 있으면 삭제
                    # (앞서 team과 name에 "[네이버] 손*리" 형태로 예약번호나 캐싱 정보를 연동했으므로 매칭하여 지웁니다)
                    cur.execute("DELETE FROM bookings WHERE name LIKE ? OR team LIKE ?", (f"%{booking_id}%", f"%{booking_id}%"))
                    
                    # C. 메일을 '읽음' 처리하여 다음 서버 구동 시 중복 스캔 방지
                    mail.store(num, '+FLAGS', '\\Seen')
                    cancel_count += 1
                    print(f"💥 [서버 구동 취소 처리] 예약번호: {booking_id} 데이터 전격 삭제 완료")

        conn.commit()
        cur.close()
        conn.close()
        mail.logout()
        print(f"🎉 [스캔 종료] 총 {cancel_count}건의 취소 데이터가 완벽히 정리되었습니다.")

    except Exception as e:
        print(f"❌ [서버 구동 취소 스캔 실패] 에러 내용: {e}")



def parse_and_save_naver_email(raw_email_content):
    """
    구글 웹훅/IMAP으로 가져온 이메일 원문(Raw Text)을 디코딩하고
    규칙에 맞게 파싱하여 DB에 INIT 상태로 저장하는 함수
    """
    try:
        decoded_body = ""

        # 1. 메일 원문에서 text/plain 또는 text/html의 base64 본문 텍스트 추출
        if "Content-Transfer-Encoding: base64" in raw_email_content:
            parts = raw_email_content.split("Content-Transfer-Encoding: base64")
            if len(parts) > 1:
                base64_snippet = parts[1].split("-------Boundary")[0].strip()
                
                try:
                    # ⭐️ 아스키 문자인지 체크 후 안전하게 디코딩 시도
                    if base64_snippet.isascii():
                        decoded_body = base64.b64decode(base64_snippet).decode('utf-8', errors='ignore')
                    else:
                        # 한글이 섞여 있는 테스트 데이터 등은 디코딩을 건너뛰고 본문 그대로 사용합니다.
                        decoded_body = base64_snippet
                except Exception:
                    decoded_body = base64_snippet
                
        
        if not decoded_body:
            decoded_body = raw_email_content

        # 2. 정밀 정규표현식(Regex) 매칭 기법 적용
        booking_id_match = re.search(r'예약번호\s*([A-Z0-9-]+|[0-9]+)', decoded_body)
        name_match = re.search(r'예약자(?:명)?\s*([가-힣A-Za-z*]+)', decoded_body)
        datetime_match = re.search(r'이용일시\s*([0-9]{4}\.[0-9]{2}\.[0-9]{2}\.[^\n]+)', decoded_body)
        room_match = re.search(r'예약상품\s*([^\n]+)', decoded_body)

        # 3. ⭐️ 데이터 조각 정제 작업 (님자 빼기 및 [네이버] 헤더 추가)
        booking_id = booking_id_match.group(1).strip() if booking_id_match else None
        
        # [교정] 메일에서 추출된 순수 이름 문자열 확보
        raw_name = name_match.group(1).strip() if name_match else "미확인"
        
        if raw_name != "미확인":
            # 💡 이름 맨 뒤의 '님' 또는 '님 ' 글자를 정규식으로 완벽 제거
            cleaned_name = re.sub(r'님\s*$', '', raw_name)
        else:
            cleaned_name = "미확인"

        final_name = cleaned_name

        raw_room = room_match.group(1).strip() if room_match else "미확인"
        room_name_extract = re.search(r'(C[1-9]|룸[0-9]|[A-Z][0-9])', raw_room)
        room_name = room_name_extract.group(1) if room_name_extract else raw_room

        # 4. 네이버 날짜/시간 포맷 정형화 (예: "2026.05.18.(월) 오후 4:00")
        use_date = datetime.now().strftime('%Y-%m-%d')
        use_time_key = "00-00"
        
        if datetime_match:
            raw_dt_text = datetime_match.group(1)
            
            # 📅 날짜 추출 (YYYY-MM-DD 형식으로 변환)
            date_part = re.search(r'([0-9]{4})\.([0-9]{2})\.([0-9]{2})', raw_dt_text)
            if date_part:
                use_date = f"{date_part.group(1)}-{date_part.group(2)}-{date_part.group(3)}"
            
            # ⏰ 시간 추출 및 대시보드 키(HH-MM) 형식으로 변환
            time_part = re.search(r'(오전|오후)\s*([0-9]{1,2}):([0-9]{2})', raw_dt_text)
            if time_part:
                ampm = time_part.group(1)
                hour = int(time_part.group(2))
                minute = time_part.group(3)
                
                if ampm == "오후" and hour != 12:
                    hour += 12
                elif ampm == "오전" and hour == 12:
                    hour = 0
                    
                use_time_key = f"{hour:02d}-{minute}"

        # 5. 필수 값 검증
        if not booking_id:
            print("⚠️ 네이버 예약 메일 형식이 아니거나 예약번호를 찾을 수 없습니다.")
            return False

        # 6. SQLite DB 테이블에 'INIT' 상태로 데이터 저장
        conn = sqlite3.connect(DB_FILE, timeout=30)
        cur = conn.cursor()
        
        # 💡 정제된 final_name 변수가 매핑되도록 쿼리 인자 수정
        cur.execute('''
            INSERT INTO naver_mail_cache 
            (booking_id, masked_name, use_date, use_time_key, room_name, status)
            VALUES (?, ?, ?, ?, ?, 'INIT')
        ''', (booking_id, final_name, use_date, use_time_key, room_name))
        
        conn.commit()
        affected_rows = cur.rowcount
        conn.close()

        if affected_rows > 0:
            print(f"🚀 [INIT 저장 완료] 번호: {booking_id} / 이름: {final_name} / 날짜: {use_date} / 시간: {use_time_key} / 룸: {room_name}")
            return {
                'booking_id': booking_id,
                'use_date': use_date,
                'use_time_key': use_time_key,
                'masked_name': final_name,
                'room_name': room_name
            }
        else:
            print(f"ℹ️ 이미 DB에 존재하는 예약 건입니다. (번호: {booking_id})")
            return False

    except Exception as e:
        print(f"❌ 메일 파싱 및 DB 저장 중 치명적 오류 발생: {e}")
        return False


load_dotenv()

def sync_missed_emails():
    GMAIL_USER = os.getenv("GMAIL_USER")
    GMAIL_APP_PASSWORD = os.getenv("GMAIL_APP_PASSWORD")
    
    if not GMAIL_USER or not GMAIL_APP_PASSWORD:
        print("❌ [오류] .env 파일에 이메일 또는 앱 비밀번호 설정이 누락되었습니다.")
        return

    print("🔄 [시스템] Gmail IMAP 서버 연결 중...")
    
    try:
        # 1. Gmail IMAP 안전 연결 및 로그인
        mail = imaplib.IMAP4_SSL("imap.gmail.com", 993)
        mail.login(GMAIL_USER, GMAIL_APP_PASSWORD)
        
        # 💡 [교정] 취소 처리 시 메일을 읽음('\Seen') 상태로 마킹해야 하므로 readonly=False로 변경합니다.
        mail.select("inbox", readonly=False)

        print("✅ [시스템] Gmail 로그인 성공! 받은편지함 통합 스캔(누락 예약 + 취소 청소)을 시작합니다.")

        # 2. 최근 2일간의 날짜 포맷 및 검색 조건 생성
        since_date = (datetime.now() - timedelta(days=2)).strftime("%d-%b-%Y")
        search_criteria = f'(SINCE "{since_date}" UNSEEN)'
        print(f"🔍 [시스템] 최근 2일간({since_date})의 메일 중 '안 읽은 새 메일' 목록을 요청합니다...")

        status, messages = mail.uid('search', None, search_criteria)

        if status == "OK":
            mail_ids = messages[0].split()
            print(f"📂 [시스템] 최근 2일간 수신된 메일 총 {len(mail_ids)}건 발견. 네이버 필터링 시작...")
            
            cancel_count = 0
            confirm_count = 0
            
            for m_id in reversed(mail_ids):
                res, msg_data = mail.uid('fetch', m_id, "(BODY[HEADER.FIELDS (SUBJECT)])")
                
                for response_part in msg_data:
                    if isinstance(response_part, tuple):
                        msg = email.message_from_bytes(response_part[1])
                        
                        # 메일 제목 디코딩 처리
                        raw_subject = msg.get('Subject', '')
                        subject, encoding = decode_header(raw_subject)[0]
                        if isinstance(subject, bytes):
                            subject = subject.decode(encoding if encoding else 'utf-8', errors='ignore')
                        
                        # 💡 [핵심 변형] '네이버'와 '예약'이 들어간 메일 중, 확정과 취소를 찢어서 판단합니다.
                        if '네이버' in subject and '예약' in subject:
                            
                            # 조건에 맞으면 본문(RFC822)을 통째로 긁어옵니다.
                            _, full_msg_data = mail.uid('fetch', m_id, "(RFC822)")
                            full_msg = email.message_from_bytes(full_msg_data[0][1])
                            
                            # 본문 텍스트 추출
                            body = ""
                            if full_msg.is_multipart():
                                for part in full_msg.walk():
                                    if part.get_content_type() == "text/plain":
                                        body = part.get_payload(decode=True).decode('utf-8', errors='ignore')
                                        break
                            else:
                                body = full_msg.get_payload(decode=True).decode('utf-8', errors='ignore')

                            # 예약번호 정규식 추출
                            import re
                            booking_id_match = re.search(r'예약번호\s*([A-Z0-9-]+|[0-9]+)', body)
                            booking_id = booking_id_match.group(1).strip() if booking_id_match else None

                            if not booking_id:
                                continue

                            conn = sqlite3.connect(DB_FILE, timeout=30)
                            cur = conn.cursor()

                            try:
                                # 🚨 Case 1: [예약 취소] 메일인 경우
                                if '취소' in subject:
                                    with conn:
                                        cur.execute("DELETE FROM naver_mail_cache WHERE booking_id = ?", (booking_id,))
                                        cur.execute("DELETE FROM bookings WHERE name LIKE ? OR team LIKE ?", (f"%{booking_id}%", f"%{booking_id}%"))
                                    
                                    mail.uid('store', m_id, '+FLAGS', '\\Seen')
                                    cancel_count += 1
                                    print(f"💥 [서버구동 청소] 취소 감지되어 예약번호 {booking_id} 관련 데이터를 모두 지웠습니다.")

                                # 🟢 Case 2: [예약 확정] 메일인 경우
                                elif '확정' in subject:
                                    # 중복 체크도 열린 커넥션 안에서 깔끔하게 처리
                                    cur.execute("SELECT 1 FROM naver_mail_cache WHERE booking_id = ?", (booking_id,))
                                    is_exist = cur.fetchone()
                                    
                                    if not is_exist:                                       
                                        # 💡 [교정]: 파티룸 문구가 있으면 캐시 상태를 'confirm'으로 저장
                                        parse_and_save_naver_email(body)
                                        if '파티룸' in body:
                                            cur.execute(
                                                """
                                                UPDATE naver_mail_cache 
                                                SET status = 'confirm' 
                                                WHERE booking_id = ?
                                                """, 
                                                (booking_id,)
                                            )
                                            conn.commit()

                                        confirm_count += 1
                                        print(f"✅ [서버구동 복구] 누락되었던 예약 확정 건 복구 완료 (번호: {booking_id})")

                                        mail.uid('store', m_id, '+FLAGS', '\\Seen')

                            except sqlite3.OperationalError as db_err:
                                web.logger.error(f"⚠️ [서버구동 스캔중 지연 발생] 새치기 요청 처리중으로 인한 대기. 에러: {db_err}")
                            finally:
                                # 🎯 [교정 2]: 디비 연결이 살아있고 열려있는 존재일 때만 '안전하게 딱 1번만' 닫아줍니다.
                                try:
                                    cur.close()
                                    conn.close()
                                except:
                                    pass # 혹시나 이미 완전히 파기된 상태라면 에러 내지 말고 부드럽게 패스!
            
            print(f"📊 [스캔 통계] 서버 오픈 준비 완료: 복구 {confirm_count}건 / 취소 청소 {cancel_count}건")
            
        mail.close()
        mail.logout()
        print("🚀 [시스템] 서버 구동 시점 통합 메일 동기화가 완전히 끝났습니다. 대시보드가 청정합니다.")
        
    except Exception as e:
        print(f"❌ [오류] 서버 구동 중 통합 메일 동기화 실패: {str(e)}")





@web.route('/api/queue/<int:qid>', methods=['PUT'])
def update_queue_item(qid):
    data = request.json or {}
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    ensure_queue_items_order_column(conn)
    cur = conn.cursor()

    cur.execute('SELECT * FROM queue_items WHERE id=?', (qid,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return jsonify({"status": "error", "message": "Queue item not found"}), 404

    room_in = data.get('room', row['room'])
    room = str(room_in).strip().upper()
    if room not in {'C1', 'C2', 'B1', 'B2'}:
        conn.close()
        return jsonify({"status": "error", "message": "Invalid room"}), 400

    name = str(data.get('name', row['name'] or '')).strip()
    team = str(data.get('team', row['team'] or '')).rstrip()[:10]
    level = str(data.get('level', row['level'] or '')).strip()
    people = str(data.get('people', row['people'] or '')).strip()
    bid = int(data.get('bid', row['bid'] or 0) or 0)
    party_room = 1 if data.get('partyRoom', bool(row['party_room'])) else 0
    has_order_no_in = 'order_no' in data
    if has_order_no_in:
        order_no = int(data.get('order_no') or 0)
    elif room != (row['room'] or ''):
        cur.execute('SELECT COALESCE(MAX(order_no), 0) FROM queue_items WHERE room=?', (room,))
        order_no = (cur.fetchone()[0] or 0) + 1
    else:
        order_no = int(row['order_no'] or 0)

    cur.execute(
        'UPDATE queue_items SET room=?, order_no=?, name=?, team=?, level=?, people=?, bid=?, party_room=? WHERE id=?',
        (room, order_no, name, team, level, people, bid, party_room, qid)
    )
    conn.commit()
    conn.close()
    # 대기리스트 변경 시 대시보드에 알림
    socketio.emit('room_or_queue_changed')
    return jsonify({"status": "success"})


@web.route('/api/queue/<int:qid>', methods=['DELETE'])
def delete_queue_item(qid):
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    cur.execute('DELETE FROM queue_items WHERE id=?', (qid,))
    conn.commit()
    conn.close()
    # 대기리스트 변경 시 대시보드에 알림
    socketio.emit('room_or_queue_changed')
    return jsonify({"status": "success"})

@web.route('/api/pad_status', methods=['GET'])
def get_pad_status():
    """
    대시보드에서 방 상태(룸 카드)를 표시하기 위한 API
    현재 게임 상태, 팀 정보, 시간 등을 반환합니다.
    """
    return jsonify({
        "map": PAD_MAP,
        "status": pad_status,
        "data": pad_data
    })


# ========================================================
# Web - Walkin
# ========================================================
@web.route('/walkin')
def walk_in_page():
    return render_template('walkin.html')


@web.route('/walkin-test')
def walk_in_test_page():
    return render_template('walkin-test.html')


@web.route('/api/walkin/add', methods=['POST'])
def add_walkin():
    data = request.json or {}
    agreed = 1 if data.get('is_agreed') else 0

    name = str(data.get('name', '')).strip()
    if not name:
        return jsonify({"success": False, "status": "error", "message": "성함을 입력해 주세요."}), 400

    # 현재 날짜와 시간 생성
    now = datetime.now()
    visit_date = now.strftime('%Y-%m-%d')  # 2026-04-23
    visit_time = now.strftime('%H:%M:%S')  # 17:59:15

    team = str(data.get('team', '')).strip()
    level = str(data.get('level', data.get('diff', ''))).strip()
    adult_count = max(int(data.get('adult_count') or 0), 0)
    child_count = max(int(data.get('child_count') or 0), 0)
    fallback_people = int(data.get('people', data.get('person', 0)) or 0)
    people = adult_count + child_count if (adult_count or child_count) else fallback_people
    room_size = str(data.get('room_size', '')).strip()
    room_fast = 1 if data.get('room_fast') else 0
    phone = str(data.get('phone', '')).strip()

    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("PRAGMA table_info(walkins)")
    columns = {row[1] for row in c.fetchall()}

    if {'level', 'people'}.issubset(columns):
        query = '''INSERT INTO walkins 
           (name, team, level, people, room_size, room_fast, initial_level, initial_room_size, initial_room_fast, adult_count, child_count, phone, is_agreed, visit_date, visit_time, status) 
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'waiting')'''
        params = (name, team, level, people, room_size, room_fast, level, room_size, room_fast, adult_count, child_count, phone, agreed, visit_date, visit_time)
    else:
        query = '''INSERT INTO walkins 
           (name, team, diff, person, room_size, room_fast, initial_level, initial_room_size, initial_room_fast, adult_count, child_count, phone, is_agreed, visit_date, visit_time, status) 
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'waiting')'''
        params = (name, team, level, people, room_size, room_fast, level, room_size, room_fast, adult_count, child_count, phone, agreed, visit_date, visit_time)

    c.execute(query, params)
    conn.commit()
    conn.close()
    # 워크인 추가 시 대시보드에 알림
    socketio.emit('walkin_added')
    return jsonify({"success": True, "status": "success"})



@web.route('/api/walkin/list')
def get_walkin_list():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("PRAGMA table_info(walkins)")
    columns = {row[1] for row in c.fetchall()}

    level_col = 'level' if 'level' in columns else 'diff'
    people_col = 'people' if 'people' in columns else 'person'
    status_filter = "WHERE status='waiting'" if 'status' in columns else ''
    adult_col = 'adult_count' if 'adult_count' in columns else '0'
    child_col = 'child_count' if 'child_count' in columns else '0'

    room_fast_col = 'room_fast' if 'room_fast' in columns else '0'

    c.execute(f'''SELECT id, name, team, {level_col}, {people_col}, room_size, {room_fast_col}, {adult_col}, {child_col}, phone 
                 FROM walkins 
                 {status_filter}
                 ORDER BY id ASC''')
    rows = c.fetchall()
    conn.close()

    result = []
    for r in rows:
        result.append({
            "id": r[0],
            "name": r[1],
            "team": r[2],
            "level": r[3],
            "people": r[4],
            "room_size": r[5],
            "room_fast": bool(r[6]),
            "adult_count": r[7] or 0,
            "child_count": r[8] or 0,
            "phone": r[9]
        })
    return jsonify(result)


@web.route('/api/walkin/complete', methods=['POST'])
def complete_walkin():
    try:
        data = request.json
        walkin_id = data.get('id')
        
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        # 해당 ID의 상태를 waiting -> entered로 변경
        c.execute("UPDATE walkins SET status='entered' WHERE id=?", (walkin_id,))
        conn.commit()
        conn.close()
        return jsonify({"status": "success"})
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500

@web.route('/api/walkin/history') # 경로를 구분해줍니다
@team_list_api_required
def get_walkin_history():
    target_date = request.args.get('date') # JS에서 보낸 날짜값
    
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    
    # 컬럼 체크 (기존 로직 활용)
    c.execute("PRAGMA table_info(walkins)")
    columns = {row[1] for row in c.fetchall()}
    
    level_col = 'level' if 'level' in columns else 'diff'
    intake_level_col = 'initial_level' if 'initial_level' in columns else level_col
    people_col = 'people' if 'people' in columns else 'person'
    adult_col = 'adult_count' if 'adult_count' in columns else '0'
    child_col = 'child_count' if 'child_count' in columns else '0'
    visit_date_col = 'visit_date' if 'visit_date' in columns else "strftime('%Y-%m-%d', 'now')"

    # 방 사이즈 관련 컬럼 체크 (room_size, room_fast)
    size_col = 'initial_room_size' if 'initial_room_size' in columns else ('room_size' if 'room_size' in columns else "''")
    fast_col = 'initial_room_fast' if 'initial_room_fast' in columns else ('room_fast' if 'room_fast' in columns else "0")

    # 쿼리: 특정 날짜의 데이터를 시간순(visit_time)으로 정렬
    # 만약 visit_time이 없다면 id 순으로 정렬합니다.
    time_sort = "visit_time ASC" if 'visit_time' in columns else "id ASC"
    
    c.execute(f'''SELECT id, name, team, {intake_level_col}, {people_col}, {adult_col}, {child_col}, phone, visit_time, status, {size_col}, {fast_col}
                 FROM walkins 
                 WHERE visit_date = ?
                 ORDER BY {time_sort}''', (target_date,))
    
    rows = c.fetchall()
    conn.close()

    result = []
    for r in rows:
        # r[10]은 room_size, r[11]은 room_fast (1 또는 0)
        base_size = r[10] if r[10] else ""
        is_fast = r[11] == 1 or r[11] == "1" or r[11] is True
        
        # 'F' 접두어 붙이기 로직
        display_label = base_size
        if is_fast and base_size:
            display_label = f"F{base_size}"
        elif is_fast and not base_size:
            display_label = "F"

        result.append({
            "id": r[0],
            "name": r[1],
            "team": r[2],
            "level": r[3],
            "people": r[4],
            "adult_count": r[5] or 0,
            "child_count": r[6] or 0,
            "phone": r[7],
            "visit_time": r[8] if r[8] else "미지정",
            "status": r[9], # 'waiting' 또는 'entered' 상태 표시용
            "room_flag_label": display_label or "-"  # JS에서 사용할 필드명
        })
    return jsonify(result)




# ========================================================
# 실행
# ========================================================
if __name__ == "__main__":
    init_db()
    init_google_sheets()

    dev_reload = os.getenv("DEV_RELOAD", "0") == "1"
    
    log_file = get_today_log_file()
    print(f"log_file: {log_file}")

    
# Flask 앱 컨텍스트가 활성화될 때 구동 함수 실행
    def run_sync_in_background():
        with web.app_context():
            sync_missed_emails()

    threading.Thread(target=run_sync_in_background, daemon=True).start()
    
    # reloader(parent/child) 환경에서 log_monitor가 중복 실행되지 않도록 가드
    should_start_monitor = (not dev_reload) or (os.getenv("WERKZEUG_RUN_MAIN") == "true")
    if should_start_monitor:
        threading.Thread(target=log_monitor, daemon=True).start()

    socketio.run(
        web, 
        host=SERVER_HOST,
        port=SERVER_PORT,
        debug=dev_reload,
        use_reloader=dev_reload
    )
