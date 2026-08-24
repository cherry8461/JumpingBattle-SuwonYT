import sqlite3
from openpyxl import load_workbook
from datetime import datetime, date, time

DB_FILE = "game_data.db"
EXCEL_FILE = "data.xlsx"   # 엑셀 파일명

print("🚀 엑셀 DB 입력 시작")

wb = load_workbook(EXCEL_FILE, data_only=True)
ws = wb.active

conn = sqlite3.connect(DB_FILE)
cur = conn.cursor()

PAD_REVERSE = {"C1": 0, "C2": 1, "B1": 2, "B2": 3}

count = 0
skip = 0

LEVEL_MAP = {
    "여름맵": "Summer",
    "산타맵": "Santa",
    "키즈맵": "Kids",
    "NORMAL": "Normal",
    "BASIC": "Basic",
    "HARD": "Hard",
    "EASY": "Easy",
    "CHALLENGER": "Challenger",
    "SPACE": "Space",
    "KIDS": "Kids",
    "SANTA": "Santa",
}

for i, row in enumerate(ws.iter_rows(min_row=1), start=1):
    try:
        d        = row[1]   # 날짜
        t        = row[2]   # 시간
        pad      = row[5]   # PAD
        map_text = row[8]   # 사이즈_레벨
        score    = row[10]  # 점수
        team     = row[13]  # 팀명

        # -------------------------
        # 날짜
        # -------------------------
        if isinstance(d.value, (date, datetime)):
            date_part = d.value.date() if isinstance(d.value, datetime) else d.value
        else:
            date_part = datetime.strptime(str(d.value).strip(), "%Y-%m-%d").date()

        # -------------------------
        # 시간
        # -------------------------
        if isinstance(t.value, (time, datetime)):
            time_part = t.value.time() if isinstance(t.value, datetime) else t.value
        else:
            time_part = datetime.strptime(str(t.value).strip(), "%H:%M:%S").time()

        dt = datetime.combine(date_part, time_part)
        time_str = dt.strftime("%Y-%m-%d %H:%M")

        # -------------------------
        # PAD
        # -------------------------
        pad_id = PAD_REVERSE.get(str(pad.value).strip())
        if pad_id is None:
            raise ValueError("PAD 인식 실패")

        raw_map = str(map_text.value).strip()

        # -------------------------
        # size / level 파싱
        # -------------------------
        if raw_map == "?":
            size = "대형"
            level = "Santa"

        elif "_" in raw_map:
            size, level = raw_map.split("_", 1)

        elif "-" in raw_map:
            size, level = raw_map.split("-", 1)

        else:
            raise ValueError(f"맵 형식 인식 실패: {raw_map}")

        # -------------------------
        # level 영어 변환
        # -------------------------
        level = LEVEL_MAP.get(level.strip(), level.strip())
        

        cur.execute("""
            INSERT OR IGNORE INTO game_records
            (time, pad_id, map_id, size, level, team, score)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            time_str,
            pad_id,
            None,
            size,
            level.replace("level-", "").capitalize(),
            str(team.value).strip(),
            int(score.value)
        ))

        if cur.rowcount > 0:
            count += 1

    except Exception as e:
        skip += 1
        print(f"⚠ 스킵 (row {i}): {e}")

conn.commit()
conn.close()

print(f"✅ 엑셀 DB 입력 완료: {count}건")
print(f"⚠ 스킵: {skip}건")
input("엔터 누르면 종료")
